"""Hoja Profesores: reporte de carga por profesor, calculado con formulas.

Un bloque por profesor: sus datos (id, nombre, grado, tope efectivo), el
detalle de lo que imparte y una fila TOTAL. El detalle no se conoce al generar
(depende de lo que se elija en Asignacion), asi que se reservan
`filas_por_profesor` lineas rellenadas con BUSCARV sobre la tabla auxiliar
CargaPorProfesor de la hoja Datos (claves '<id>#<n>'). Son formulas planas: se
comportan igual en Excel y en Calc. La ultima linea reservada avisa '(+N más)'
si el profesor imparte mas cosas de las que caben en el bloque.
"""
from openpyxl.utils import quote_sheetname

from comun import formato, leyenda
from departamento import estilos
from departamento import layout as L
from departamento.hoja_asignacion import NOMBRE_HOJA as HOJA_ASIGNACION
from departamento.modelo import Departamento

NOMBRE_HOJA = "Profesores"

# Columnas del detalle: BUSCARV sobre CargaPorProfesor (clave en la col. 1).
# A=asignatura (col 2), B=tipo (3), C=grupo (4), D=horas (5).
_COLS_DETALLE = (("A", 2), ("B", 3), ("C", 4), ("D", 5))


def construir_hoja_profesores(wb, depto: Departamento) -> None:
    ws = wb.create_sheet(NOMBRE_HOJA)
    ws[f"A{L.FILA_TITULO}"] = f"Profesores — {depto.nombre} — {depto.semestre}"
    ws[f"A{L.FILA_TITULO}"].font = estilos.fuente_encabezado()

    fpp = depto.filas_por_profesor
    for i, p in enumerate(depto.profesores):
        _construir_bloque(ws, depto, i, p)

    ultimo_total = L.prof_fila_total(len(depto.profesores) - 1, fpp)
    # Padding aproximado para Calc: filas mas altas en toda la zona de bloques.
    formato.aplicar_alto_filas(ws, L.PROF_FILA_PRIMER_BLOQUE, ultimo_total,
                               estilos.ALTO_FILA)
    _escribir_leyenda(ws, ultimo_total + 2)
    ws.freeze_panes = f"A{L.PROF_FILA_PRIMER_BLOQUE}"
    formato.autoajustar_columnas(ws, extra=4)
    # La columna A del detalle muestra nombres de asignatura por formula
    # (autoajustar la ignora): el ancho se fija con los nombres posibles.
    formato.fijar_ancho_por_textos(
        ws, "A", [a.nombre for a in depto.asignaturas] + ["Asignatura"], extra=4)


def _rango_asignacion(col: str, n_filas: int) -> str:
    hoja = quote_sheetname(HOJA_ASIGNACION)
    return (f"{hoja}!${col}${L.FILA_PRIMERA_CARGA}"
            f":${col}${L.fila_carga(n_filas - 1)}")


def _construir_bloque(ws, depto: Departamento, idx: int, profesor) -> None:
    fpp = depto.filas_por_profesor
    n_filas = len(depto.filas())
    rango_prof = _rango_asignacion(L.COL_PROFESOR, n_filas)
    conteo = f'COUNTIF({rango_prof},"{profesor.id}")'

    # Cabecera y valores del profesor.
    fila_cab = L.prof_fila_cabecera(idx, fpp)
    _fila_encabezado(ws, fila_cab, ("Id", "Nombre", "Grado", "Tope horas"))
    fila_val = L.prof_fila_valores(idx, fpp)
    tope = depto.tope_efectivo(profesor)
    for col, valor in zip("ABCD", (profesor.id, profesor.nombre, profesor.grado,
                                   tope if tope is not None else "")):
        ws[f"{col}{fila_val}"] = valor

    # Subcabecera y detalle reservado.
    _fila_encabezado(ws, L.prof_fila_subcabecera(idx, fpp),
                     ("Asignatura", "Tipo", "Grupo", "Horas"))
    for k in range(fpp):
        fila = L.prof_fila_detalle(idx, k, fpp)
        for col, col_tabla in _COLS_DETALLE:
            buscar = (f'IFERROR(VLOOKUP("{profesor.id}#{k + 1}",'
                      f'CargaPorProfesor,{col_tabla},0),"")')
            if col == "A" and k == fpp - 1:
                # Ultima linea reservada: si hay desborde, avisa cuantas faltan
                # (la propia linea deja de mostrarse, por eso se suma 1).
                buscar = (f'IF({conteo}>{fpp},'
                          f'"(+"&({conteo}-{fpp}+1)&" más)",{buscar})')
            ws[f"{col}{fila}"] = f"={buscar}"

    # TOTAL de horas del profesor y alerta de sobrecarga.
    fila_total = L.prof_fila_total(idx, fpp)
    ws[f"A{fila_total}"] = "TOTAL"
    ws[f"A{fila_total}"].font = estilos.fuente_encabezado()
    ws[f"D{fila_total}"] = (f'=SUMIF({rango_prof},"{profesor.id}",'
                            f"{_rango_asignacion(L.COL_HORAS, n_filas)})")
    if tope is not None:
        # Rojo en la fila TOTAL cuando el acumulado supera el tope del bloque.
        ws.conditional_formatting.add(
            f"A{fila_total}:D{fila_total}",
            estilos.regla_formula(f"$D${fila_total}>$D${fila_val}",
                                  estilos.COLOR_SOBRECARGA))

    rango = f"A{fila_cab}:D{fila_total}"
    formato.aplicar_borde_tabla(ws, rango, interno=estilos.lado_fino(),
                                externo=estilos.lado_medio())
    formato.aplicar_alineacion(ws, rango, estilos.alineacion_padding())


def _fila_encabezado(ws, fila: int, textos) -> None:
    for i, texto in enumerate(textos):
        celda = ws.cell(row=fila, column=i + 1, value=texto)
        celda.font = estilos.fuente_encabezado()
        celda.fill = estilos.fill(estilos.COLOR_ENCABEZADO)


def _escribir_leyenda(ws, fila: int) -> None:
    leyenda.escribir_leyenda(ws, f"A{fila}", (
        (estilos.COLOR_SOBRECARGA, "Profesor por encima de su tope de horas"),
    ))
