"""Hoja oculta Datos del generador del departamento.

Contiene las tablas de apoyo de las demas hojas:
- A:D  tabla de profesores (id, nombre, grado, tope efectivo), con los rangos
       nombrados ProfesoresValidos (ids, para el desplegable) y ProfesoresTabla
       (para BUSCARV de nombre y tope).
- F:J  tabla auxiliar de carga por profesor: una fila por fila de carga de la
       hoja Asignacion. La columna F construye la clave '<prof>#<n>' con un
       CONTAR.SI de rango creciente (n = numero de aparicion del profesor hasta
       esa fila); G:J llevan asignatura, tipo, grupo y horas. El rango nombrado
       CargaPorProfesor permite a la hoja Profesores rellenar su detalle con
       BUSCARV planos, sin formulas matriciales.
"""
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate

from departamento import layout as L
from departamento.modelo import Departamento

NOMBRE_HOJA = "Datos"
NOMBRE_HOJA_ASIGNACION = "Asignación"


def _rango_nombrado(nombre: str, celda_ini: str, celda_fin: str) -> DefinedName:
    ref = (f"{quote_sheetname(NOMBRE_HOJA)}!"
           f"{absolute_coordinate(celda_ini)}:{absolute_coordinate(celda_fin)}")
    return DefinedName(nombre, attr_text=ref)


def _formula_clave(idx: int) -> str:
    """Formula de la clave '<prof>#<n>' para la fila de carga `idx`. El CONTAR.SI
    va de la primera fila de carga hasta la propia (rango creciente), de modo que
    cuenta cuantas veces ha aparecido ese profesor hasta aqui."""
    hoja = quote_sheetname(NOMBRE_HOJA_ASIGNACION)
    celda = f"{hoja}!${L.COL_PROFESOR}${L.fila_carga(idx)}"
    rango = (f"{hoja}!${L.COL_PROFESOR}${L.FILA_PRIMERA_CARGA}"
             f":${L.COL_PROFESOR}${L.fila_carga(idx)}")
    return f'=IF({celda}="","",{celda}&"#"&COUNTIF({rango},{celda}))'


def construir_hoja_datos(wb, depto: Departamento) -> None:
    ws = wb.create_sheet(NOMBRE_HOJA)
    ws.sheet_state = "hidden"

    # Tabla de profesores, sin encabezado (BUSCARV directo). El tope escrito es
    # el efectivo (propio o global), para que las hojas no repitan la regla.
    for i, p in enumerate(depto.profesores, start=1):
        ws[f"A{i}"] = p.id
        ws[f"B{i}"] = p.nombre
        ws[f"C{i}"] = p.grado
        tope = depto.tope_efectivo(p)
        ws[f"D{i}"] = tope if tope is not None else ""
    n_prof = len(depto.profesores)
    wb.defined_names.add(_rango_nombrado("ProfesoresValidos", "A1", f"A{n_prof}"))
    wb.defined_names.add(_rango_nombrado("ProfesoresTabla", "A1", f"D{n_prof}"))

    # Tabla auxiliar de carga: clave por formula, datos estaticos de cada fila.
    filas = depto.filas()
    for i, f in enumerate(filas):
        r = i + 1
        ws[f"F{r}"] = _formula_clave(i)
        ws[f"G{r}"] = f.asignatura.nombre
        ws[f"H{r}"] = f.tipo
        ws[f"I{r}"] = f.grupo if f.grupo is not None else "-"
        ws[f"J{r}"] = f.horas
    if filas:
        wb.defined_names.add(_rango_nombrado("CargaPorProfesor", "F1", f"J{len(filas)}"))
