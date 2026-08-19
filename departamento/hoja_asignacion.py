"""Hoja Asignacion: la unica editable del libro del departamento.

Una fila por fila de carga (Conf o grupo de CP). El usuario elige el profesor
en la columna F con un desplegable de ayuda no bloqueante; el nombre completo
aparece al lado por formula. Los colores avisan de filas sin profesor
(amarillo) y de ids fuera de la lista (ambar).
"""
from openpyxl.worksheet.datavalidation import DataValidation

from comun import formato, leyenda
from departamento import estilos
from departamento import layout as L
from departamento.modelo import Departamento

NOMBRE_HOJA = "Asignación"

ALTO_FILA_CARGA = 22  # puntos: parte del padding aproximado para Calc


def construir_hoja_asignacion(wb, depto: Departamento) -> None:
    ws = wb.create_sheet(NOMBRE_HOJA)
    filas = depto.filas()

    ws[f"A{L.FILA_TITULO}"] = f"Asignación — {depto.nombre} — {depto.semestre}"
    ws[f"A{L.FILA_TITULO}"].font = estilos.fuente_encabezado()

    _escribir_encabezados(ws)
    _escribir_filas(ws, filas)
    _aplicar_dropdown(ws, len(filas))
    _aplicar_formato_condicional(ws, len(filas))
    _aplicar_formato(ws, depto, len(filas))
    _escribir_leyenda(ws, len(filas))

    # Encabezados fijos al hacer scroll: todo lo anterior a la primera carga.
    ws.freeze_panes = f"A{L.FILA_PRIMERA_CARGA}"


def _escribir_encabezados(ws) -> None:
    fila = L.FILA_ENCABEZADO_ASIGNACION
    for i, texto in enumerate(L.ENCABEZADOS_ASIGNACION):
        celda = ws.cell(row=fila, column=i + 1, value=texto)
        celda.font = estilos.fuente_encabezado()
        celda.fill = estilos.fill(estilos.COLOR_ENCABEZADO)


def _escribir_filas(ws, filas) -> None:
    for i, f in enumerate(filas):
        r = L.fila_carga(i)
        ws[f"{L.COL_ASIGNATURA}{r}"] = f.asignatura.nombre
        ws[f"{L.COL_CARRERA}{r}"] = f.asignatura.carrera
        ws[f"{L.COL_TIPO}{r}"] = f.tipo
        ws[f"{L.COL_GRUPO}{r}"] = f.grupo if f.grupo is not None else "-"
        ws[f"{L.COL_HORAS}{r}"] = f.horas
        # F queda vacia: ahi se decide el profesor. G muestra su nombre.
        ws[f"{L.COL_NOMBRE}{r}"] = (
            f'=IF(F{r}="","",IFERROR(VLOOKUP(F{r},ProfesoresTabla,2,0),'
            f'"(desconocido)"))')


def _aplicar_dropdown(ws, n_filas: int) -> None:
    # Fuente: rango nombrado 'ProfesoresValidos' de la hoja Datos. Sin '=' inicial
    # (openpyxl escribe formula1 verbatim) y con aviso 'information' no bloqueante,
    # para poder escribir un id nuevo sin que Calc/Excel lo rechacen.
    dv = DataValidation(type="list", formula1="ProfesoresValidos", allow_blank=True,
                        showErrorMessage=True, errorStyle="information")
    ws.add_data_validation(dv)
    dv.sqref = (f"{L.COL_PROFESOR}{L.FILA_PRIMERA_CARGA}"
                f":{L.COL_PROFESOR}{L.fila_carga(n_filas - 1)}")


def _aplicar_formato_condicional(ws, n_filas: int) -> None:
    fila_ini = L.FILA_PRIMERA_CARGA
    rango = f"A{fila_ini}:{L.COL_ULTIMA}{L.fila_carga(n_filas - 1)}"
    # Columna $F fijada y fila relativa: toda la fila evalua su propia celda de
    # profesor. Amarillo: sin profesor. Ambar: id fuera de la lista.
    ws.conditional_formatting.add(
        rango, estilos.regla_formula(f'$F{fila_ini}=""', estilos.COLOR_SIN_PROFESOR))
    ws.conditional_formatting.add(
        rango,
        estilos.regla_formula(
            f'AND($F{fila_ini}<>"",COUNTIF(ProfesoresValidos,$F{fila_ini})=0)',
            estilos.COLOR_PROFESOR_DESCONOCIDO))


def _aplicar_formato(ws, depto: Departamento, n_filas: int) -> None:
    fila_fin = L.fila_carga(n_filas - 1)
    rango = f"A{L.FILA_ENCABEZADO_ASIGNACION}:{L.COL_ULTIMA}{fila_fin}"
    formato.aplicar_borde_tabla(ws, rango, interno=estilos.lado_fino(),
                                externo=estilos.lado_medio())
    # Padding aproximado para Calc: sangria + centrado vertical + filas mas altas.
    formato.aplicar_alineacion(ws, rango, estilos.alineacion_padding())
    formato.aplicar_alto_filas(ws, L.FILA_PRIMERA_CARGA, fila_fin, ALTO_FILA_CARGA)
    formato.autoajustar_columnas(ws, extra=4)
    # La columna Nombre muestra el resultado de una formula (autoajustar la
    # ignora): su ancho se fija con los nombres que puede llegar a mostrar.
    formato.fijar_ancho_por_textos(
        ws, L.COL_NOMBRE,
        [p.nombre for p in depto.profesores] + ["(desconocido)"], extra=4)


def _escribir_leyenda(ws, n_filas: int) -> None:
    fila = L.fila_carga(n_filas - 1) + 2
    leyenda.escribir_leyenda(ws, f"A{fila}", (
        (estilos.COLOR_SIN_PROFESOR, "Falta asignar profesor"),
        (estilos.COLOR_PROFESOR_DESCONOCIDO, "Profesor fuera de la lista"),
    ))
