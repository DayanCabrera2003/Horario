"""Hoja Asignaturas: reporte de quien imparte cada asignatura.

Un bloque por asignatura con sus filas de carga (Conf y cada grupo de CP) y el
profesor elegido en cada una, por referencia directa a la hoja Asignacion (las
filas de una asignatura son contiguas alli, en el mismo orden). El titulo del
bloque se pinta verde cuando todas sus filas tienen profesor y naranja mientras
falte alguna.
"""
from openpyxl.utils import quote_sheetname

from comun import formato, leyenda
from departamento import estilos
from departamento import layout as L
from departamento.hoja_asignacion import NOMBRE_HOJA as HOJA_ASIGNACION
from departamento.modelo import Departamento, filas_de_carga

NOMBRE_HOJA = "Asignaturas"

SUBCABECERA = ("Tipo", "Grupo", "Horas", "Profesor", "Nombre")
COL_ULTIMA = "E"


def construir_hoja_asignaturas(wb, depto: Departamento) -> None:
    ws = wb.create_sheet(NOMBRE_HOJA)
    ws[f"A{L.FILA_TITULO}"] = f"Asignaturas — {depto.nombre} — {depto.semestre}"
    ws[f"A{L.FILA_TITULO}"].font = estilos.fuente_encabezado()

    alturas = 0      # acumulado de alturas de los bloques anteriores
    idx_carga = 0    # indice global de fila de carga (posicion en Asignacion)
    for a in depto.asignaturas:
        n = len(filas_de_carga((a,)))
        _construir_bloque(ws, a, alturas, idx_carga, n)
        alturas += L.altura_bloque_asignatura(n)
        idx_carga += n

    # La ultima fila ocupada es la anterior al blanco del ultimo bloque.
    _escribir_leyenda(ws, L.asig_fila_titulo(alturas) + 1)
    ws.freeze_panes = f"A{L.ASIG_FILA_PRIMER_BLOQUE}"
    formato.autoajustar_columnas(ws, extra=4)
    # Las columnas Profesor y Nombre muestran resultados de formulas
    # (autoajustar las ignora): el ancho se fija con los valores posibles.
    formato.fijar_ancho_por_textos(ws, "D", [p.id for p in depto.profesores],
                                   extra=4)
    formato.fijar_ancho_por_textos(
        ws, "E", [p.nombre for p in depto.profesores] + ["(desconocido)"],
        extra=4)


def _construir_bloque(ws, asignatura, alturas: int, idx_carga: int, n: int) -> None:
    hoja = quote_sheetname(HOJA_ASIGNACION)
    fila_titulo = L.asig_fila_titulo(alturas)
    ws[f"A{fila_titulo}"] = f"{asignatura.nombre} — {asignatura.carrera}"
    ws[f"A{fila_titulo}"].font = estilos.fuente_encabezado()

    fila_sub = L.asig_fila_subcabecera(alturas)
    for i, texto in enumerate(SUBCABECERA):
        celda = ws.cell(row=fila_sub, column=i + 1, value=texto)
        celda.font = estilos.fuente_encabezado()
        celda.fill = estilos.fill(estilos.COLOR_ENCABEZADO)

    filas = filas_de_carga((asignatura,))
    for k, f in enumerate(filas):
        fila = L.asig_fila_carga(alturas, k)
        r_asig = L.fila_carga(idx_carga + k)   # fila gemela en Asignacion
        ws[f"A{fila}"] = f.tipo
        ws[f"B{fila}"] = f.grupo if f.grupo is not None else "-"
        ws[f"C{fila}"] = f.horas
        # La celda de profesor puede estar vacia y una referencia directa a una
        # celda vacia se muestra como 0; el IF la deja en blanco en ese caso.
        celda_prof = f"{hoja}!{L.COL_PROFESOR}{r_asig}"
        ws[f"D{fila}"] = f'=IF({celda_prof}="","",{celda_prof})'
        ws[f"E{fila}"] = f"={hoja}!{L.COL_NOMBRE}{r_asig}"

    # Titulo verde si ninguna fila de la asignatura esta sin profesor, naranja
    # si falta alguna. El rango evaluado son sus filas gemelas de Asignacion.
    rango_prof = (f"{hoja}!${L.COL_PROFESOR}${L.fila_carga(idx_carga)}"
                  f":${L.COL_PROFESOR}${L.fila_carga(idx_carga + n - 1)}")
    rango_titulo = f"A{fila_titulo}:{COL_ULTIMA}{fila_titulo}"
    ws.conditional_formatting.add(
        rango_titulo,
        estilos.regla_formula(f"COUNTBLANK({rango_prof})=0", estilos.COLOR_COMPLETA))
    ws.conditional_formatting.add(
        rango_titulo,
        estilos.regla_formula(f"COUNTBLANK({rango_prof})>0", estilos.COLOR_INCOMPLETA))

    rango = f"A{fila_titulo}:{COL_ULTIMA}{L.asig_fila_carga(alturas, n - 1)}"
    formato.aplicar_borde_tabla(ws, rango, interno=estilos.lado_fino(),
                                externo=estilos.lado_medio())
    formato.aplicar_alineacion(ws, rango, estilos.alineacion_padding())


def _escribir_leyenda(ws, fila: int) -> None:
    leyenda.escribir_leyenda(ws, f"A{fila}", (
        (estilos.COLOR_COMPLETA, "Asignatura completa (todo asignado)"),
        (estilos.COLOR_INCOMPLETA, "Asignatura incompleta (faltan profesores)"),
    ))
