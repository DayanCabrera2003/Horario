"""Utilidades para aplicar formato de presentacion (bordes, ancho de columnas)
sobre una worksheet ya construida. Separado de `estilos`, que solo define objetos
de estilo puros, y de `layout`, que solo calcula posiciones."""
from openpyxl.styles import Border, Side


def aplicar_borde(ws, rango: str, borde: Border) -> None:
    """Aplica `borde` a cada celda del `rango` (p. ej. "C3:D15").

    openpyxl no tiene un borde "de rango": el borde es un estilo por celda, asi que
    se recorre el rango y se asigna a cada una.
    """
    for fila in ws[rango]:
        for celda in fila:
            celda.border = borde


def aplicar_borde_tabla(ws, rango: str, interno: Side, externo: Side) -> None:
    """Bordea un rango rectangular: `externo` en las caras del perimetro y
    `interno` en las caras internas.

    openpyxl no tiene borde de rango; el borde es un estilo por celda. Se
    compone el Border de cada celda segun su posicion en el rango: las celdas
    del borde reciben el lado `externo` en la cara que da al exterior.
    """
    filas = ws[rango]
    n_filas = len(filas)
    n_cols = len(filas[0]) if n_filas else 0
    for i, fila in enumerate(filas):
        for j, celda in enumerate(fila):
            celda.border = Border(
                top=externo if i == 0 else interno,
                bottom=externo if i == n_filas - 1 else interno,
                left=externo if j == 0 else interno,
                right=externo if j == n_cols - 1 else interno,
            )


def aplicar_borde_inferior(ws, rango: str, lado: Side) -> None:
    """Sobrescribe solo la cara inferior de cada celda del `rango` con `lado`,
    conservando las otras tres caras del borde existente.

    Se usa para trazar separadores horizontales (p. ej. la linea gruesa entre
    turnos) encima de un enrejado ya bordeado, sin borrar el resto del borde.
    """
    for fila in ws[rango]:
        for celda in fila:
            b = celda.border
            celda.border = Border(left=b.left, right=b.right, top=b.top, bottom=lado)


def aplicar_estilo_encabezado(ws, celdas, fuente, relleno) -> None:
    """Aplica `fuente` y `relleno` a las celdas indicadas (iterable de
    coordenadas). Se estilizan celdas concretas, no rangos, para no pintar
    celdas en blanco."""
    for coord in celdas:
        ws[coord].font = fuente
        ws[coord].fill = relleno


def aplicar_relleno(ws, rango: str, relleno) -> None:
    """Aplica `relleno` (PatternFill) a cada celda del `rango`. Para pintar el
    fondo de una franja de celdas (p. ej. las filas donde van las aulas)."""
    for fila in ws[rango]:
        for celda in fila:
            celda.fill = relleno


def aplicar_alineacion(ws, rango: str, alineacion) -> None:
    """Aplica `alineacion` a cada celda del `rango`."""
    for fila in ws[rango]:
        for celda in fila:
            celda.alignment = alineacion


def aplicar_ajuste_texto(ws, rango: str, alineacion) -> None:
    """Aplica `alineacion` (con wrap_text) a cada celda del `rango`, para que
    el contenido largo salte de linea en vez de desbordarse."""
    aplicar_alineacion(ws, rango, alineacion)


def aplicar_alto_filas(ws, fila_ini: int, fila_fin: int, alto: float) -> None:
    """Fija el alto (en puntos) de las filas de `fila_ini` a `fila_fin` inclusive,
    para dar aire vertical (parte del padding aproximado)."""
    for fila in range(fila_ini, fila_fin + 1):
        ws.row_dimensions[fila].height = alto


def autoajustar_columnas(ws, min_ancho: int = 8, max_ancho: int = 45,
                         extra: int = 2) -> None:
    """Fija el ancho de cada columna al del texto mas largo que contiene.

    openpyxl no tiene "autofit" real (requiere renderizar), asi que se estima por
    longitud de caracteres. Las formulas ("=...") se ignoran: su texto no coincide con
    el valor que se muestra e inflaria la columna. El ancho se limita a [min_ancho, max_ancho].
    """
    largos: dict[str, int] = {}
    for fila in ws.iter_rows():
        for celda in fila:
            valor = celda.value
            if valor is None:
                continue
            if isinstance(valor, str) and valor.startswith("="):
                continue
            largo = len(str(valor))
            if largo > largos.get(celda.column_letter, 0):
                largos[celda.column_letter] = largo
    for letra, largo in largos.items():
        ws.column_dimensions[letra].width = min(max_ancho, max(min_ancho, largo + extra))
