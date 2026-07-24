"""Primitivas de estilo genericas (sin datos de dominio) reutilizables por
cualquier generador. Los colores propios de cada dominio viven en el `estilos`
de su paquete; aqui solo hay fabricas de objetos de estilo de openpyxl."""
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule


def fill(hex_rgb: str) -> PatternFill:
    return PatternFill(start_color=hex_rgb, end_color=hex_rgb, fill_type="solid")


def regla_formula(formula: str, color_hex: str) -> FormulaRule:
    return FormulaRule(formula=[formula], fill=fill(color_hex))


def lado_fino() -> Side:
    """Lado de grosor fino, para el enrejado interno de las tablas."""
    return Side(style="thin")


def lado_medio() -> Side:
    """Lado de grosor medio, para el borde exterior (perimetro) de las tablas."""
    return Side(style="medium")


def borde_fino() -> Border:
    """Borde fino en las cuatro caras, para delimitar las celdas de las tablas."""
    lado = lado_fino()
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def alineacion_ajuste() -> Alignment:
    """Alineacion con ajuste de texto (wrap): el contenido que no cabe salta de
    linea y la fila crece, para que nada quede cortado."""
    return Alignment(wrap_text=True, vertical="top")


def fuente_encabezado() -> Font:
    """Fuente en negrita para las celdas de encabezado."""
    return Font(bold=True)
