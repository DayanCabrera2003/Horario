from openpyxl import Workbook

from departamento.modelo import Profesor, Asignatura, Departamento
from departamento.hoja_datos import construir_hoja_datos


def _departamento():
    return Departamento(
        nombre="Matemática Aplicada", semestre="2026-2027 / 1",
        tope_horas=160, filas_por_profesor=10,
        profesores=(
            Profesor(id="PIAD", nombre="Pedro I. Alonso", grado="Dr."),
            Profesor(id="MARA", nombre="Maria Ramirez", grado="MSc.", tope_horas=80),
        ),
        asignaturas=(
            Asignatura(id="EST-CC", nombre="Estadística (CC)",
                       carrera="Ciencia de la Computación",
                       horas_conf=32, horas_cp=32, grupos_cp=2),
        ),
    )


def _hoja():
    wb = Workbook()
    wb.remove(wb.active)
    construir_hoja_datos(wb, _departamento())
    return wb, wb["Datos"]


def test_hoja_oculta():
    _, ws = _hoja()
    assert ws.sheet_state == "hidden"


def test_tabla_profesores_con_tope_efectivo():
    _, ws = _hoja()
    assert [ws[f"A{i}"].value for i in (1, 2)] == ["PIAD", "MARA"]
    assert ws["B1"].value == "Pedro I. Alonso"
    assert ws["C2"].value == "MSc."
    # Tope efectivo: el global para PIAD, el propio para MARA.
    assert ws["D1"].value == 160
    assert ws["D2"].value == 80


def test_rangos_nombrados():
    wb, _ = _hoja()
    nombres = wb.defined_names
    assert "ProfesoresValidos" in nombres
    assert "ProfesoresTabla" in nombres
    assert "CargaPorProfesor" in nombres
    assert "$A$1:$A$2" in nombres["ProfesoresValidos"].attr_text
    assert "$A$1:$D$2" in nombres["ProfesoresTabla"].attr_text
    # 3 filas de carga (Conf + 2 grupos de CP) -> F1:J3.
    assert "$F$1:$J$3" in nombres["CargaPorProfesor"].attr_text


def test_tabla_auxiliar_clave_y_datos():
    _, ws = _hoja()
    # La clave de la fila 1 apunta a la fila 4 de Asignacion y numera las
    # apariciones del profesor con un CONTAR.SI de rango creciente.
    clave = ws["F1"].value
    assert clave.startswith("=IF(")
    assert "Asignación'!$F$4" in clave
    assert "COUNTIF('Asignación'!$F$4:$F$4" in clave
    # La fila 3 cierra el rango creciente en $F$6.
    assert "COUNTIF('Asignación'!$F$4:$F$6" in ws["F3"].value
    # Datos estaticos de la fila de carga: asignatura, tipo, grupo, horas.
    assert ws["G1"].value == "Estadística (CC)"
    assert ws["H1"].value == "Conf"
    assert ws["I1"].value == "-"
    assert ws["J1"].value == 32
    assert ws["H2"].value == "CP"
    assert ws["I2"].value == 1
