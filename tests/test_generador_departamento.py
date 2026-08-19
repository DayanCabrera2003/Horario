from openpyxl import load_workbook

from departamento.generador import generar


CONFIG = """
departamento:
  nombre: Matemática Aplicada
  semestre: "2026-2027 / 1"
  tope_horas: 160
profesores:
  - {id: PIAD, nombre: "Pedro I. Alonso", grado: "Dr."}
asignaturas:
  - {id: EST-CC, nombre: "Estadística (CC)", carrera: "CC",
     horas_conf: 32, horas_cp: 32, grupos_cp: 2}
"""


def test_genera_libro_completo(tmp_path):
    cfg = tmp_path / "departamento.yaml"
    cfg.write_text(CONFIG, encoding="utf-8")
    salida = tmp_path / "gestion.xlsx"

    ruta = generar(config_path=cfg, salida=salida)

    assert ruta == salida and salida.exists()
    wb = load_workbook(salida)
    # Asignacion primera y activa; Datos oculta al final.
    assert wb.sheetnames == ["Asignación", "Profesores", "Asignaturas", "Datos"]
    assert wb.active.title == "Asignación"
    assert wb["Datos"].sheet_state == "hidden"
