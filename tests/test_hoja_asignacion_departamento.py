from openpyxl import Workbook

from departamento.modelo import Profesor, Asignatura, Departamento
from departamento.hoja_datos import construir_hoja_datos
from departamento.hoja_asignacion import construir_hoja_asignacion
from departamento import estilos


def _departamento():
    return Departamento(
        nombre="Matemática Aplicada", semestre="2026-2027 / 1",
        tope_horas=160, filas_por_profesor=10,
        profesores=(
            Profesor(id="PIAD", nombre="Pedro I. Alonso", grado="Dr."),
            Profesor(id="MARA", nombre="Maria Ramirez", grado="MSc."),
        ),
        asignaturas=(
            Asignatura(id="EST-CC", nombre="Estadística (CC)",
                       carrera="Ciencia de la Computación",
                       horas_conf=32, horas_cp=32, grupos_cp=2),
            Asignatura(id="EST-MAT", nombre="Estadística (Mat)",
                       carrera="Matemática", horas_conf=32, horas_cp=32,
                       grupos_cp=1),
        ),
    )


def _hoja():
    wb = Workbook()
    wb.remove(wb.active)
    depto = _departamento()
    construir_hoja_datos(wb, depto)
    construir_hoja_asignacion(wb, depto)
    return wb["Asignación"]


def test_titulo_y_encabezados():
    ws = _hoja()
    assert "Matemática Aplicada" in ws["A1"].value
    assert "2026-2027 / 1" in ws["A1"].value
    assert [c.value for c in ws["A3":"G3"][0]] == [
        "Asignatura", "Carrera", "Tipo", "Grupo", "Horas", "Profesor", "Nombre"]


def test_filas_de_carga():
    ws = _hoja()
    # 5 filas: EST-CC (Conf, CP-1, CP-2) + EST-MAT (Conf, CP-1).
    assert ws["A4"].value == "Estadística (CC)"
    assert [ws[f"C{r}"].value for r in range(4, 9)] == [
        "Conf", "CP", "CP", "Conf", "CP"]
    assert ws["D4"].value == "-"
    assert ws["D6"].value == 2
    assert ws["E4"].value == 32
    assert ws["A8"].value == "Estadística (Mat)"
    # La celda de profesor queda vacia (la decision se toma en el Excel).
    assert ws["F4"].value is None


def test_formula_nombre_profesor():
    ws = _hoja()
    formula = ws["G4"].value
    assert formula.startswith("=IF(F4=")
    assert "VLOOKUP(F4,ProfesoresTabla,2,0)" in formula


def test_desplegable_no_bloqueante():
    ws = _hoja()
    dvs = ws.data_validations.dataValidation
    assert len(dvs) == 1
    dv = dvs[0]
    assert dv.formula1 == "ProfesoresValidos"
    assert dv.errorStyle == "information"
    assert "F4:F8" in str(dv.sqref)


def test_formato_condicional_fila_completa():
    ws = _hoja()
    reglas = {}
    for rango, lista in ws.conditional_formatting._cf_rules.items():
        for regla in lista:
            color = regla.dxf.fill.start_color.rgb[-6:]
            reglas[color] = (str(rango.sqref), regla.formula[0])
    sq_amarillo, f_amarillo = reglas[estilos.COLOR_SIN_PROFESOR]
    assert sq_amarillo == "A4:G8"
    assert f_amarillo == '$F4=""'
    sq_ambar, f_ambar = reglas[estilos.COLOR_PROFESOR_DESCONOCIDO]
    assert sq_ambar == "A4:G8"
    assert "COUNTIF(ProfesoresValidos,$F4)=0" in f_ambar


def test_leyenda_y_paneles():
    ws = _hoja()
    assert ws.freeze_panes == "A4"
    textos = [ws[f"B{r}"].value for r in range(10, 13)]
    assert any("profesor" in (t or "").lower() for t in textos)
