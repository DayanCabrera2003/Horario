from openpyxl import Workbook
from horarios import leyenda, estilos


def test_escribir_leyenda_pone_color_y_texto_por_fila():
    wb = Workbook()
    ws = wb.active
    items = [("FFF176", "Aula invalida"), ("FFB74D", "Asignatura desconocida")]
    leyenda.escribir_leyenda(ws, "B2", items)
    # Primera fila: B2 coloreada, C2 con el texto
    assert ws["B2"].fill.fgColor.rgb.endswith("FFF176")
    assert ws["C2"].value == "Aula invalida"
    # Segunda fila una mas abajo
    assert ws["B3"].fill.fgColor.rgb.endswith("FFB74D")
    assert ws["C3"].value == "Asignatura desconocida"
