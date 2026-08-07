from openpyxl import Workbook
from tribunales.modelo import (Profesor, Estudiante, Local, Momento, Dia, Tesis, Facultad)
from tribunales.hoja_localizar import construir_hoja_localizar, NOMBRE_HOJA, COL_ROL


def _fac():
    return Facultad(
        profesores=(Profesor("PIAD", "P", "Dr."),),
        estudiantes=(Estudiante("JPER", "Juan"),),
        locales=(Local("POST", "Postgrado"),),
        dias=(Dia("2026-07-27", (Momento("09:00", "10:00"),)),),
        tesis=(Tesis("JPER", "PIAD", "PIAD", "PIAD", "PIAD"),),
    )


def test_celda_entrada_y_titulo():
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_localizar(wb, _fac())
    ws = wb[NOMBRE_HOJA]
    assert ws["A1"].value is not None            # etiqueta "Localizar a:"
    textos = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Postgrado" in t and "2026-07-27" in t for t in textos)  # titulo dia-local


def test_regla_participacion_referencia_hoja_dia():
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_localizar(wb, _fac())
    ws = wb[NOMBRE_HOJA]
    formulas = []
    for _sqref, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if getattr(rule, "formula", None):
                formulas.extend(rule.formula)
    # Debe comparar la entrada global contra celdas de la hoja del dia.
    assert any("$B$1" in f and "2026-07-27" in f for f in formulas)


def test_columna_rol_muestra_el_rol():
    # Junto a cada momento, una columna Rol con una formula que resuelve en calidad
    # de que participa la persona buscada, comparando contra las columnas de la hoja
    # del dia (B..F: estudiante, tutor, oponente, presidente, secretario).
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_localizar(wb, _fac())
    ws = wb[NOMBRE_HOJA]
    textos = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert "Rol" in textos                       # cabecera de la columna de rol
    formulas = [ws[f"{COL_ROL}{r}"].value for r in range(1, ws.max_row + 1)
                if isinstance(ws[f"{COL_ROL}{r}"].value, str)
                and ws[f"{COL_ROL}{r}"].value.startswith("=")]
    assert formulas, "esperaba una formula de rol en la columna B"
    f = formulas[0]
    for etiqueta in ("Estudiante", "Tutor", "Oponente", "Presidente", "Secretario"):
        assert etiqueta in f
    assert "2026-07-27" in f                      # referencia a la hoja del dia


def test_localizar_tiene_leyenda():
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_localizar(wb, _fac())
    ws = wb[NOMBRE_HOJA]
    textos = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Leyenda" in t for t in textos)
