from openpyxl import Workbook
from horarios.modelo import Grupo, Anio, Facultad
from horarios.hoja_datos import construir_hoja_datos, _formula_firma_anio, NOMBRE_HOJA


def _fac():
    g = [Grupo("C", 1, 1, 1), Grupo("M", 1, 1, 1)]
    anios = {"C1": Anio("C", 1, ()), "M1": Anio("M", 1, ())}
    return Facultad(aulas=("Aula 1", "Lab"), dias=("Lunes",), turnos=2, grupos=tuple(g), anios=anios)


def test_crea_hoja_oculta_y_lista_aulas():
    fac = _fac(); wb = Workbook()
    construir_hoja_datos(wb, fac)
    ws = wb[NOMBRE_HOJA]
    assert ws.sheet_state == "hidden"
    valores = [c.value for col in ws.iter_cols() for c in col]
    assert "Aula 1" in valores and "Lab" in valores


def test_rango_nombrado_aulas_validas():
    fac = _fac(); wb = Workbook()
    construir_hoja_datos(wb, fac)
    assert "AulasValidas" in wb.defined_names


def test_firma_usa_countif():
    fac = _fac(); wb = Workbook()
    celdas = construir_hoja_datos(wb, fac)
    # celdas: {(dia, turno, aula): direccion_de_la_firma}
    dir_firma = celdas[("Lunes", 1, "Aula 1")]   # p.ej. "Datos!C1" (referencia calificada)
    ws = wb[NOMBRE_HOJA]
    coord_local = dir_firma.split("!")[1]         # indexar la hoja con la coord local, no "Datos!C1"
    assert ws[coord_local].value.startswith("=")
    assert "COUNTIF" in ws[coord_local].value


def test_firma_referencia_ambos_anios_y_mix():
    # Con dos años distintos presentes, la fórmula debe referenciar las celdas-aula
    # de ambos grupos, incluir ambos códigos de año y la rama de fallback "MIX".
    fac = _fac(); wb = Workbook()
    celdas = construir_hoja_datos(wb, fac)
    ws = wb[NOMBRE_HOJA]
    coord_local = celdas[("Lunes", 1, "Aula 1")].split("!")[1]
    formula = ws[coord_local].value
    assert "C111" in formula and "M111" in formula     # celdas-aula de ambos grupos
    assert '"C1"' in formula and '"M1"' in formula      # ambos códigos de año
    assert '"MIX"' in formula                           # rama de conflicto


def test_formula_firma_anio_directo():
    # Prueba unitaria directa del helper para un montaje de 2 años.
    anios_por_codigo = {
        "C1": [Grupo("C", 1, 1, 1)],
        "M1": [Grupo("M", 1, 1, 1)],
    }
    formula = _formula_firma_anio(anios_por_codigo, 0, 1, "Aula 1")
    assert formula.startswith("=")
    assert "COUNTIF" in formula
    assert '"C1"' in formula and '"M1"' in formula
    assert '"MIX"' in formula


def test_mapa_cubre_todas_las_combinaciones():
    fac = _fac(); wb = Workbook()
    celdas = construir_hoja_datos(wb, fac)
    assert len(celdas) == len(fac.dias) * fac.turnos * len(fac.aulas)
    # y todas las direcciones son únicas
    assert len(set(celdas.values())) == len(celdas)
