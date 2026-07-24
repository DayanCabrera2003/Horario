from openpyxl import load_workbook
from tribunales.generador import generar


def _escribir(tmp_path):
    cfg = tmp_path / "tribunal.yaml"
    cfg.write_text("""
profesores:
  - {id: PIAD, nombre: "Pedro", grado: "Dr."}
  - {id: MARA, nombre: "Maria", grado: "MSc."}
  - {id: LGOM, nombre: "Luis", grado: "Dr."}
  - {id: ANSU, nombre: "Ana", grado: "MSc."}
estudiantes:
  - {id: JPER, nombre: "Juan"}
locales:
  - {id: POST, nombre: "Postgrado"}
dias:
  - fecha: 2026-07-27
    momentos:
      - {inicio: "09:00", fin: "10:00"}
tesis:
  - {estudiante: JPER, tutor: PIAD, oponente: MARA, presidente: LGOM, secretario: ANSU}
""", encoding="utf-8")
    return cfg


def test_genera_hojas_esperadas(tmp_path):
    cfg = _escribir(tmp_path)
    salida = tmp_path / "tesis.xlsx"
    generar(config_path=cfg, asignaciones_path=None, salida=salida)
    wb = load_workbook(salida)
    assert "Localizar" in wb.sheetnames
    assert "2026-07-27" in wb.sheetnames
    assert "Datos" in wb.sheetnames


def test_genera_con_asignaciones(tmp_path):
    cfg = _escribir(tmp_path)
    asig = tmp_path / "asig.yaml"
    asig.write_text('- {estudiante: JPER, local: POST, fecha: 2026-07-27, momento: "09:00-10:00"}\n',
                    encoding="utf-8")
    salida = tmp_path / "tesis.xlsx"
    generar(config_path=cfg, asignaciones_path=asig, salida=salida)
    wb = load_workbook(salida)
    ws = wb["2026-07-27"]
    assert ws["B3"].value == "JPER"   # columna Estudiante, primer momento, primer local
