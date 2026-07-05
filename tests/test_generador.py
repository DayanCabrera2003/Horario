"""Tests de integración para horarios.generador."""
from openpyxl import load_workbook
from horarios.config import cargar_facultad
from horarios.generador import generar
from horarios import layout as L

BASE = """
aulas: [Aula 1, Lab]
dias: [Lunes, Martes]
turnos: 6
carreras:
  C:
    nombre: Carrera C
    años:
      1:
        sesiones: { 1: { grupos: [1, 2] } }
        asignaturas:
          - { id: L-C, nombre: "Lógica", frecuencia: 1 }
"""


def _config(tmp_path):
    cfg = tmp_path / "facultad.yaml"
    cfg.write_text(BASE, encoding="utf-8")
    return cfg


# --- Test dado en la especificación ---

def test_genera_workbook_con_hojas(tmp_path):
    cfg = _config(tmp_path)
    salida = tmp_path / "out.xlsx"
    generar(config_path=cfg, horarios_path=None, salida=salida)
    wb = load_workbook(salida)
    assert "Aulas" in wb.sheetnames
    assert "C111" in wb.sheetnames and "C112" in wb.sheetnames
    assert "Datos" in wb.sheetnames
    assert wb["Datos"].sheet_state == "hidden"


# --- Test extra 1: ruta con horarios (faithful-copy) ---

def test_horarios_yaml_rellena_celdas(tmp_path):
    """generar con horarios_path escribe asig y aula en la hoja del grupo."""
    cfg = _config(tmp_path)
    horarios_yaml = tmp_path / "horarios.yaml"
    horarios_yaml.write_text(
        "C111:\n  Lunes:\n    1:\n      asig: L-C\n      aula: Aula 1\n",
        encoding="utf-8",
    )
    salida = tmp_path / "out.xlsx"
    generar(config_path=cfg, horarios_path=horarios_yaml, salida=salida)

    wb = load_workbook(salida)
    ws = wb["C111"]
    # día 0 = Lunes, turno 1
    assert ws[L.celda_asig(0, 1)].value == "L-C"
    assert ws[L.celda_aula(0, 1)].value == "Aula 1"
    # Hoja C112 (sin horario) debe tener esa misma celda vacía
    assert wb["C112"][L.celda_asig(0, 1)].value is None


# --- Test extra 2: humo de fórmulas en modo esqueleto ---

def test_formulas_presentes_en_modo_esqueleto(tmp_path):
    """En modo esqueleto las hojas de grupo contienen COUNTIF y Aulas tiene SUBSTITUTE."""
    cfg = _config(tmp_path)
    salida = tmp_path / "out.xlsx"
    generar(config_path=cfg, horarios_path=None, salida=salida)

    wb = load_workbook(salida)

    # Hoja de grupo: debe contener al menos un =COUNTIF( en la columna Asignadas
    ws_grupo = wb["C111"]
    formulas_grupo = [
        c.value
        for row in ws_grupo.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert any("COUNTIF(" in f for f in formulas_grupo), (
        f"No se encontró =COUNTIF en C111; fórmulas vistas: {formulas_grupo[:10]}"
    )

    # Hoja Aulas: debe contener al menos un =SUBSTITUTE(
    ws_aulas = wb["Aulas"]
    formulas_aulas = [
        c.value
        for row in ws_aulas.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert any("SUBSTITUTE" in f for f in formulas_aulas), (
        f"No se encontró SUBSTITUTE en Aulas; fórmulas vistas: {formulas_aulas[:10]}"
    )
