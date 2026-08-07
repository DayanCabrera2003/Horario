from openpyxl import Workbook
from horarios.modelo import Grupo, Asignatura, Anio, Facultad, Horario, Asignacion
from horarios.hoja_grupo import construir_hoja_grupo
from horarios import layout as L
from horarios import estilos


def _facultad():
    anio = Anio(carrera="C", numero=1, asignaturas=(
        Asignatura("L-C", "Lógica", 1),
        Asignatura("Pro-C", "Programación", 2),
    ))
    g = Grupo("C", 1, 1, 1)
    return Facultad(aulas=("Aula 1", "Lab"), dias=("Lunes", "Martes"),
                    turnos=6, grupos=(g,), anios={"C1": anio}), g


def test_escribe_id_y_formulas():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    assert ws[L.CELDA_GRUPO_ID].value == "C111"
    # Asignadas de la primera asignatura = COUNTIF sobre el rango del horario
    asignadas = ws[L.celda_asig_tabla_asignadas(0)].value
    assert asignadas.startswith("=COUNTIF(")
    # fixture: 2 días, 6 turnos -> rango del horario B4:C15
    assert "B4:C15" in asignadas
    faltan = ws[L.celda_asig_tabla_faltan(0)].value
    assert faltan.startswith("=")


def test_hay_dropdowns_de_aula_y_asignatura():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    # dos validaciones de lista: aula (filas de aula) y asignatura (filas de asignatura)
    dvs = ws.data_validations.dataValidation
    assert len(dvs) == 2
    formulas = {dv.formula1 for dv in dvs}
    assert "AulasValidas" in formulas
    assert "$I$4:$I$5" in formulas
    # openpyxl escribe formula1 verbatim: un '=' inicial rompe el dropdown en Excel.
    assert all(not f.startswith("=") for f in formulas)


def test_dropdowns_no_bloquean_valores_fuera_de_lista():
    # En Calc, un dropdown de lista sin accion de error configurada rechaza (Stop) los
    # valores nuevos. Para poder escribir aulas/asignaturas fuera de la lista sin que se
    # bloquee, ambas validaciones usan errorStyle 'information' (aviso no bloqueante) con
    # el mensaje de error activo. El desplegable se mantiene como ayuda.
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    dvs = ws.data_validations.dataValidation
    assert dvs, "esperaba validaciones de datos"
    for dv in dvs:
        assert dv.errorStyle == "information"
        assert dv.showErrorMessage is True


def test_tablas_tienen_bordes():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    # Esquina superior izquierda de la tabla de asignaturas: perimetro medium
    assert ws["I3"].border.left.style == "medium"
    assert ws["I3"].border.top.style == "medium"
    # Cara interna de la rejilla de horario: fina
    assert ws[L.celda_asig(0, 1)].border.top.style == "thin"
    # Etiqueta de turno (esquina superior de la columna A): perimetro medium
    assert ws[f"A{L.fila_asig(1)}"].border.left.style == "medium"


def test_fondo_en_filas_de_aula():
    # Las filas de aula llevan un fondo neutro que las marca como zona de aulas;
    # las filas de asignatura no reciben ese fondo.
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    aula = ws[L.celda_aula(0, 1)].fill
    assert estilos.COLOR_FONDO_AULA in (aula.fgColor.rgb or "")
    asig = ws[L.celda_asig(0, 1)].fill
    assert estilos.COLOR_FONDO_AULA not in (asig.fgColor.rgb or "")


def test_padding_en_celdas_del_horario():
    # Padding aproximado para Calc: sangria + centrado vertical en la rejilla y
    # mayor alto de fila. El .xlsx no tiene padding real de celda.
    from horarios.hoja_grupo import ALTO_FILA_HORARIO
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    alin = ws[L.celda_asig(0, 1)].alignment
    assert alin.indent >= 1
    assert alin.vertical == "center"
    # La fila de la rejilla tiene un alto mayor que el por defecto.
    assert ws.row_dimensions[L.fila_asig(1)].height == ALTO_FILA_HORARIO


def test_linea_gruesa_separa_turnos():
    # Un turno ocupa 2 filas (asignatura + aula). La frontera inferior de la fila
    # de aula de cada turno (salvo el ultimo) lleva un borde grueso separador.
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    # Turno 1: bajo su fila de aula, borde grueso (tanto en dias como en la col A).
    assert ws[L.celda_aula(0, 1)].border.bottom.style == "thick"
    assert ws[f"A{L.fila_aula(1)}"].border.bottom.style == "thick"
    # Ultimo turno (6): su borde inferior es el perimetro medio, no el separador.
    assert ws[L.celda_aula(0, fac.turnos)].border.bottom.style == "medium"


def test_frecuencia_exacta_colorea_fila_completa():
    # La regla de frecuencia exacta (verde) y la de sobre-planificada (rojo) cubren
    # la fila entera de la asignatura (I..M), no solo la columna Asignadas, y usan
    # columnas fijadas ($L/$K) con fila relativa.
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    pares = [(str(rng.sqref), r.formula[0])
             for rng in ws.conditional_formatting
             for r in ws.conditional_formatting[rng]]
    exacta = [(rng, f) for rng, f in pares if f == "$L4=$K4"]
    sobre = [(rng, f) for rng, f in pares if f == "$L4>$K4"]
    assert exacta, pares
    assert sobre, pares
    # fixture: 2 asignaturas -> filas 4 y 5, columnas I..M
    assert exacta[0][0].startswith("I4:M")
    assert sobre[0][0].startswith("I4:M")


def test_inmoviliza_dias_y_turnos():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    # Freeze en la primera celda editable: fija la fila de dias (3) y la columna A
    assert ws.freeze_panes == L.celda_asig(0, 1)  # "B4"


def test_encabezados_con_estilo():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    # Encabezado de dia (fila 3) en negrita
    assert ws[f"{L.col_dia(0)}{L.FILA_ENCABEZADO_DIAS}"].font.bold is True
    # Cabecera de la tabla de asignaturas en negrita
    assert ws["I3"].font.bold is True
    # Etiqueta de turno en negrita
    assert ws[f"A{L.fila_asig(1)}"].font.bold is True
    # Una fila de aula intercalada (en blanco) NO recibe relleno de encabezado
    fill_aula = ws[f"A{L.fila_aula(1)}"].fill
    assert fill_aula.fgColor.rgb in (None, "00000000")


def test_columnas_ajustadas_al_contenido():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    # La columna de nombres (J) debe caber la asignatura mas larga ("Programación").
    assert ws.column_dimensions["J"].width >= len("Programación") + 2


def test_hoja_grupo_tiene_leyenda():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    textos = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Leyenda" in t for t in textos)
    assert any("fuera del listado" in t for t in textos)  # descripcion de aula invalida


def test_horario_rellena_celdas():
    fac, g = _facultad()
    h = Horario(grupo_id="C111")
    h.celdas[("Lunes", 1)] = Asignacion(asig="L-C", aula="Aula 1")
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=h)
    assert ws[L.celda_asig(0, 1)].value == "L-C"
    assert ws[L.celda_aula(0, 1)].value == "Aula 1"


def test_esqueleto_deja_celdas_vacias():
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    assert ws[L.celda_asig(0, 1)].value is None


def test_regla_asig_desconocida_usa_rango_absoluto():
    # La regla "asignatura desconocida" se aplica sobre un sqref multi-rango que abarca
    # todas las filas de asignatura. El COUNTIF debe apuntar SIEMPRE a la tabla de ids
    # (rango absoluto $I$), o Excel lo desplaza por fila y comprueba el rango equivocado.
    fac, g = _facultad()
    wb = Workbook()
    ws = wb.active
    construir_hoja_grupo(ws, g, fac, horario=None)
    formulas = [r.formula[0]
                for rng in ws.conditional_formatting
                for r in ws.conditional_formatting[rng]]
    desconocida = [f for f in formulas if "COUNTIF($I$" in f]
    assert desconocida, f"esperaba un COUNTIF con rango absoluto, vi: {formulas}"
