from openpyxl import Workbook

from departamento.modelo import Profesor, Asignatura, Departamento
from departamento.hoja_datos import construir_hoja_datos
from departamento.hoja_asignacion import construir_hoja_asignacion
from departamento.hoja_asignaturas import construir_hoja_asignaturas
from departamento import estilos


def _departamento():
    return Departamento(
        nombre="Matemática Aplicada", semestre="2026-2027 / 1",
        tope_horas=160, filas_por_profesor=10,
        profesores=(Profesor(id="PIAD", nombre="Pedro I. Alonso", grado="Dr."),),
        asignaturas=(
            Asignatura(id="EST-CC", nombre="Estadística (CC)",
                       carrera="Ciencia de la Computación",
                       horas_conf=32, horas_cp=32, grupos_cp=2),
            Asignatura(id="EST-MAT", nombre="Estadística (Mat)",
                       carrera="Matemática", horas_conf=32, horas_cp=32,
                       grupos_cp=1),
        ),
    )


def _hoja():
    wb = Workbook()
    wb.remove(wb.active)
    depto = _departamento()
    construir_hoja_datos(wb, depto)
    construir_hoja_asignacion(wb, depto)
    construir_hoja_asignaturas(wb, depto)
    return wb["Asignaturas"]


def test_bloques_por_asignatura():
    ws = _hoja()
    # EST-CC: titulo en 3, subcabecera en 4, filas 5..7 (Conf + 2 CP).
    assert "Estadística (CC)" in ws["A3"].value
    assert "Ciencia de la Computación" in ws["A3"].value
    assert [c.value for c in ws["A4":"E4"][0]] == [
        "Tipo", "Grupo", "Horas", "Profesor", "Nombre"]
    assert [ws[f"A{r}"].value for r in (5, 6, 7)] == ["Conf", "CP", "CP"]
    assert ws["B6"].value == 1
    assert ws["C5"].value == 32
    # EST-MAT: bloque siguiente (altura previa 6) -> titulo en 9.
    assert "Estadística (Mat)" in ws["A9"].value
    assert [ws[f"A{r}"].value for r in (11, 12)] == ["Conf", "CP"]


def test_profesor_referencia_a_asignacion():
    ws = _hoja()
    # La fila Conf de EST-CC es la fila de carga 4 de Asignacion.
    assert ws["D5"].value == "='Asignación'!F4"
    assert ws["E5"].value == "='Asignación'!G4"
    # La fila CP de EST-MAT es la fila de carga 8.
    assert ws["D12"].value == "='Asignación'!F8"


def test_titulo_coloreado_por_completitud():
    ws = _hoja()
    reglas = []
    for rango, lista in ws.conditional_formatting._cf_rules.items():
        for regla in lista:
            reglas.append((str(rango.sqref), regla.formula[0],
                           regla.dxf.fill.start_color.rgb[-6:]))
    # Verde si ninguna fila de EST-CC (F4:F6 de Asignacion) esta en blanco.
    assert ("A3:E3", "COUNTBLANK('Asignación'!$F$4:$F$6)=0",
            estilos.COLOR_COMPLETA) in reglas
    assert ("A3:E3", "COUNTBLANK('Asignación'!$F$4:$F$6)>0",
            estilos.COLOR_INCOMPLETA) in reglas
    # EST-MAT evalua sus propias filas (F7:F8).
    assert ("A9:E9", "COUNTBLANK('Asignación'!$F$7:$F$8)=0",
            estilos.COLOR_COMPLETA) in reglas


def test_leyenda():
    ws = _hoja()
    # Ultimo bloque termina en la fila 12; leyenda dos filas despues.
    textos = [(ws[f"B{r}"].value or "") for r in (14, 15)]
    assert any("completa" in t.lower() for t in textos)
