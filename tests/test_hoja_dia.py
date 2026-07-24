from openpyxl import Workbook
from tribunales.modelo import (Profesor, Estudiante, Local, Momento, Dia, Tesis, Facultad)
from tribunales.hoja_datos import construir_hoja_datos
from tribunales.hoja_dia import construir_hoja_dia


def _fac():
    return Facultad(
        profesores=(Profesor("PIAD", "P", "Dr."), Profesor("MARA", "M", "MSc."),
                    Profesor("LGOM", "L", "Dr."), Profesor("ANSU", "A", "MSc.")),
        estudiantes=(Estudiante("JPER", "Juan"),),
        locales=(Local("POST", "Postgrado"), Local("DECA", "Decanato")),
        dias=(Dia("2026-07-27", (Momento("09:00", "10:00"), Momento("10:00", "11:00"))),),
        tesis=(Tesis("JPER", "PIAD", "MARA", "LGOM", "ANSU"),),
    )


def _hoja(fac, asigs=()):
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_datos(wb, fac)
    dia = fac.dias[0]
    ws = wb.create_sheet(dia.fecha)
    construir_hoja_dia(ws, dia, fac, asignaciones=asigs)
    return ws


def test_titulos_de_local_presentes():
    ws = _hoja(_fac())
    textos = [c.value for row in ws.iter_rows() for c in row]
    assert "Postgrado" in textos and "Decanato" in textos


def test_encabezados_de_columna():
    ws = _hoja(_fac())
    fila2 = [ws[f"{c}2"].value for c in "ABCDEF"]
    assert fila2 == ["Momento", "Estudiante", "Tutor", "Oponente", "Presidente", "Secretario"]


def test_columnas_profesor_son_lookup():
    ws = _hoja(_fac())
    # Fila del primer momento del primer local (layout: fila 3).
    assert isinstance(ws["C3"].value, str) and ws["C3"].value.startswith("=")
    assert "TesisTribunal" in ws["C3"].value


def test_dropdown_estudiantes_presente():
    ws = _hoja(_fac())
    formulas_dv = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert "EstudiantesValidos" in formulas_dv


def test_regla_colision_emitida():
    ws = _hoja(_fac())
    formulas = []
    for _sqref, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if getattr(rule, "formula", None):
                formulas.extend(rule.formula)
    # La formula de colision cuenta locales con COUNTIF y compara > 1.
    assert any("COUNTIF" in f and ">1" in f.replace(" ", "") for f in formulas)


def test_encabezado_en_negrita():
    ws = _hoja(_fac())
    assert ws["A1"].font.bold is True   # titulo de local
    assert ws["A2"].font.bold is True   # encabezado de columnas


def test_congela_columna_a():
    ws = _hoja(_fac())
    assert ws.freeze_panes == "B1"


def test_hoja_dia_tiene_leyenda_de_colision():
    ws = _hoja(_fac())
    textos = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Leyenda" in t for t in textos)
    assert any("olisi" in t for t in textos)  # "colision"/"Colision"
