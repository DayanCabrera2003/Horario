from openpyxl import Workbook
from tribunales.modelo import (Profesor, Estudiante, Local, Momento, Dia, Tesis,
                               Facultad)
from tribunales.hoja_tribunales import construir_hoja_tribunales, NOMBRE_HOJA


def _fac():
    return Facultad(
        profesores=(
            Profesor("PIAD", "Pedro", "Dr."),
            Profesor("MARA", "Maria", "MSc."),
            Profesor("LGOM", "Luis", "Dr."),
            Profesor("ANSU", "Ana", "MSc."),
        ),
        estudiantes=(Estudiante("JPER", "Juan Perez"),),
        locales=(Local("POST", "Postgrado"),),
        dias=(Dia("2026-07-27", (Momento("09:00", "10:00"),)),),
        tesis=(Tesis(estudiantes=("JPER",), tutores=("PIAD",), oponente="MARA",
                     presidente="LGOM", secretario="ANSU"),),
    )


def _construir():
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_tribunales(wb, _fac())
    return wb[NOMBRE_HOJA]


def test_encabezados():
    ws = _construir()
    assert ws["A1"].value == "Estudiante (id)"
    assert ws["B1"].value == "Estudiante"
    assert [ws[f"{c}1"].value for c in "CDEF"] == ["Tutor", "Oponente",
                                                   "Presidente", "Secretario"]
    assert ws["A1"].font.bold is True


def test_fila_con_nombres_completos():
    ws = _construir()
    # El id del estudiante se conserva (para casar con los desplegables) y el resto
    # se muestra por nombre, con el grado del profesor.
    assert ws["A2"].value == "JPER"
    assert ws["B2"].value == "Juan Perez"
    assert ws["C2"].value == "Dr. Pedro"       # tutor
    assert ws["D2"].value == "MSc. Maria"      # oponente
    assert ws["E2"].value == "Dr. Luis"        # presidente
    assert ws["F2"].value == "MSc. Ana"        # secretario


def test_id_desconocido_se_muestra_como_respaldo():
    # Si un rol referencia un id que no esta en el listado, se muestra el id tal cual.
    fac = _fac()
    tesis = (Tesis(estudiantes=("JPER",), tutores=("XXXX",), oponente="MARA",
                   presidente="LGOM", secretario="ANSU"),)
    fac = Facultad(fac.profesores, fac.estudiantes, fac.locales, fac.dias, tesis)
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_tribunales(wb, fac)
    ws = wb[NOMBRE_HOJA]
    assert ws["C2"].value == "XXXX"


def test_conjunta_y_cotutoria_unen_nombres_y_vocal():
    # Tesis conjunta (dos estudiantes) con co-tutoria (dos tutores) y vocal: los
    # nombres se unen con ' / ' y el vocal aparece en su columna.
    fac = Facultad(
        profesores=(Profesor("PIAD", "Pedro", "Dr."), Profesor("MARA", "Maria", "MSc."),
                    Profesor("LGOM", "Luis", "Dr."), Profesor("ANSU", "Ana", "MSc.")),
        estudiantes=(Estudiante("JPER", "Juan Perez"), Estudiante("MGOM", "Mario Gomez")),
        locales=(Local("POST", "Postgrado"),),
        dias=(Dia("2026-07-27", (Momento("09:00", "10:00"),)),),
        tesis=(Tesis(estudiantes=("JPER", "MGOM"), tutores=("PIAD", "MARA"),
                     oponente="LGOM", presidente="ANSU", secretario="PIAD",
                     vocal="MARA"),),
    )
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_tribunales(wb, fac)
    ws = wb[NOMBRE_HOJA]
    assert ws["B2"].value == "Juan Perez / Mario Gomez"
    assert ws["C2"].value == "Dr. Pedro / MSc. Maria"
    assert ws["G2"].value == "MSc. Maria"   # vocal


def test_inmoviliza_encabezado():
    ws = _construir()
    assert ws.freeze_panes == "A2"
