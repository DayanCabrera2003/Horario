from openpyxl import Workbook
from horarios import estilos, formato


def test_aplicar_borde_pone_borde_en_cada_celda_del_rango():
    wb = Workbook()
    ws = wb.active
    formato.aplicar_borde(ws, "B2:C3", estilos.borde_fino())
    for coord in ("B2", "C2", "B3", "C3"):
        assert ws[coord].border.left.style == "thin"


def test_aplicar_borde_no_afecta_celdas_fuera_del_rango():
    wb = Workbook()
    ws = wb.active
    formato.aplicar_borde(ws, "B2:C3", estilos.borde_fino())
    assert ws["A1"].border.left.style is None


def test_autoajustar_columnas_ensancha_segun_contenido():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B1"] = "una cadena bastante larga en B"
    formato.autoajustar_columnas(ws)
    assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width


def test_autoajustar_columnas_respeta_los_limites():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ab"                     # muy corto -> se aplica el minimo
    ws["B1"] = "z" * 200               # muy largo -> se recorta al maximo
    formato.autoajustar_columnas(ws, min_ancho=8, max_ancho=45)
    assert ws.column_dimensions["A"].width == 8
    assert ws.column_dimensions["B"].width == 45


def test_aplicar_borde_tabla_perimetro_medium_interior_thin():
    wb = Workbook()
    ws = wb.active
    formato.aplicar_borde_tabla(ws, "B2:D4",
                                estilos.lado_fino(), estilos.lado_medio())
    # Esquina superior izquierda: arriba e izquierda medium
    assert ws["B2"].border.top.style == "medium"
    assert ws["B2"].border.left.style == "medium"
    # Celda central: las cuatro caras thin
    assert ws["C3"].border.top.style == "thin"
    assert ws["C3"].border.left.style == "thin"
    assert ws["C3"].border.right.style == "thin"
    assert ws["C3"].border.bottom.style == "thin"
    # Esquina inferior derecha: abajo y derecha medium
    assert ws["D4"].border.bottom.style == "medium"
    assert ws["D4"].border.right.style == "medium"


def test_autoajustar_columnas_ignora_formulas():
    # El texto de una formula ("=...") no refleja el valor mostrado; no debe inflar el ancho.
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=COUNTIF(C4:D15,I4)+SUBSTITUTE(x)"   # formula larga
    ancho_antes = ws.column_dimensions["A"].width     # ancho por defecto
    formato.autoajustar_columnas(ws, min_ancho=8, max_ancho=45)
    # La columna solo tiene una formula: su ancho no cambia (no se cuenta el texto de la formula).
    assert ws.column_dimensions["A"].width == ancho_antes
