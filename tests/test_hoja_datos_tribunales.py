from openpyxl import Workbook
from tribunales.modelo import Estudiante, Tesis, Facultad, Profesor
from tribunales.hoja_datos import construir_hoja_datos, NOMBRE_HOJA


def _fac():
    return Facultad(
        profesores=(Profesor("PIAD", "P", "Dr."),),
        estudiantes=(Estudiante("JPER", "Juan"),),
        locales=(),
        dias=(),
        tesis=(Tesis(estudiantes=("JPER",), tutores=("PIAD",), oponente="PIAD",
                     presidente="PIAD", secretario="PIAD"),),
    )


def test_hoja_datos_oculta_y_con_tabla_tesis():
    wb = Workbook()
    wb.remove(wb.active)
    construir_hoja_datos(wb, _fac())
    ws = wb[NOMBRE_HOJA]
    assert ws.sheet_state == "hidden"
    assert ws["A1"].value == "JPER"   # estudiante de la primera tesis
    assert ws["B1"].value == "PIAD"   # tutor


def test_tesis_conjunta_una_fila_por_estudiante():
    # Una tesis conjunta (dos estudiantes) genera una fila por estudiante, ambas
    # con el mismo tribunal; los co-tutores se unen con ' / ' y el vocal va en F.
    fac = Facultad(
        profesores=(Profesor("PIAD", "P", "Dr."), Profesor("MARA", "M", "MSc.")),
        estudiantes=(Estudiante("JPER", "Juan"), Estudiante("MGOM", "Mario")),
        locales=(), dias=(),
        tesis=(Tesis(estudiantes=("JPER", "MGOM"), tutores=("PIAD", "MARA"),
                     oponente="MARA", presidente="PIAD", secretario="MARA",
                     vocal="PIAD"),),
    )
    wb = Workbook(); wb.remove(wb.active)
    construir_hoja_datos(wb, fac)
    ws = wb[NOMBRE_HOJA]
    assert ws["A1"].value == "JPER" and ws["A2"].value == "MGOM"
    assert ws["B1"].value == "PIAD / MARA" and ws["B2"].value == "PIAD / MARA"
    assert ws["F1"].value == "PIAD"   # vocal


def test_rangos_nombrados_definidos():
    wb = Workbook()
    wb.remove(wb.active)
    construir_hoja_datos(wb, _fac())
    assert "TesisTribunal" in wb.defined_names
    assert "EstudiantesValidos" in wb.defined_names
