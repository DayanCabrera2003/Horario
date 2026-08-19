from openpyxl import Workbook

from departamento.modelo import Profesor, Asignatura, Departamento
from departamento.hoja_datos import construir_hoja_datos
from departamento.hoja_asignacion import construir_hoja_asignacion
from departamento.hoja_profesores import construir_hoja_profesores
from departamento import estilos


def _departamento(tope_global=160):
    return Departamento(
        nombre="Matemática Aplicada", semestre="2026-2027 / 1",
        tope_horas=tope_global, filas_por_profesor=4,
        profesores=(
            Profesor(id="PIAD", nombre="Pedro I. Alonso", grado="Dr."),
            Profesor(id="MARA", nombre="Maria Ramirez", grado="MSc.", tope_horas=80),
        ),
        asignaturas=(
            Asignatura(id="EST-CC", nombre="Estadística (CC)",
                       carrera="Ciencia de la Computación",
                       horas_conf=32, horas_cp=32, grupos_cp=2),
        ),
    )


def _hoja(depto=None):
    wb = Workbook()
    wb.remove(wb.active)
    depto = depto or _departamento()
    construir_hoja_datos(wb, depto)
    construir_hoja_asignacion(wb, depto)
    construir_hoja_profesores(wb, depto)
    return wb["Profesores"]


def test_bloque_cabecera():
    ws = _hoja()
    # filas_por_profesor=4 -> bloques de altura 9; PIAD en 3, MARA en 12.
    assert [c.value for c in ws["A3":"D3"][0]] == ["Id", "Nombre", "Grado", "Tope horas"]
    assert ws["A4"].value == "PIAD"
    assert ws["B4"].value == "Pedro I. Alonso"
    assert ws["D4"].value == 160     # tope efectivo (global)
    assert ws["A12"].value == "Id"
    assert ws["A13"].value == "MARA"
    assert ws["D13"].value == 80     # tope propio


def test_detalle_por_buscarv():
    ws = _hoja()
    # Subcabecera en 5; detalle en 6..9.
    assert [c.value for c in ws["A5":"D5"][0]] == ["Asignatura", "Tipo", "Grupo", "Horas"]
    f = ws["A6"].value
    assert 'VLOOKUP("PIAD#1",CargaPorProfesor,2,0)' in f
    assert f.startswith("=IFERROR(")
    assert 'VLOOKUP("PIAD#2",CargaPorProfesor,3,0)' in ws["B7"].value
    assert 'CargaPorProfesor,5,0' in ws["D6"].value


def test_ultima_fila_reservada_avisa_desborde():
    ws = _hoja()
    # La fila reservada 4 (fila 9) avisa si el profesor tiene mas filas de las
    # que caben en el bloque.
    f = ws["A9"].value
    assert "más)" in f
    assert 'COUNTIF(' in f and '"PIAD"' in f


def test_total_con_sumif():
    ws = _hoja()
    assert ws["A10"].value == "TOTAL"
    f = ws["D10"].value
    # 3 filas de carga -> rango F4:F6 de Asignacion.
    assert f == ('=SUMIF(\'Asignación\'!$F$4:$F$6,"PIAD",'
                 "'Asignación'!$E$4:$E$6)")


def test_alto_de_filas_para_padding():
    ws = _hoja()
    # Bloques de la fila 3 a la ultima fila TOTAL (19): filas mas altas.
    assert ws.row_dimensions[3].height == estilos.ALTO_FILA
    assert ws.row_dimensions[19].height == estilos.ALTO_FILA


def test_alerta_sobrecarga_solo_con_tope():
    ws = _hoja()
    rangos = {str(r.sqref) for r in ws.conditional_formatting._cf_rules}
    # Regla roja sobre la fila TOTAL de cada bloque.
    assert "A10:D10" in rangos
    assert "A19:D19" in rangos

    # Sin tope global ni propio no se agrega la regla del bloque.
    depto = _departamento(tope_global=None)
    depto = Departamento(**{**depto.__dict__,
                            "profesores": (Profesor(id="PIAD", nombre="P", grado="Dr."),)})
    ws2 = _hoja(depto)
    assert not list(ws2.conditional_formatting._cf_rules)
