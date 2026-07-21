from openpyxl import Workbook
from horarios.modelo import Grupo, Anio, Facultad
from horarios.hoja_datos import construir_hoja_datos
from horarios.hoja_aulas import construir_hoja_aulas, NOMBRE_HOJA


def _fac():
    g = [Grupo("C", 1, 1, 1), Grupo("C", 1, 1, 2)]
    anios = {"C1": Anio("C", 1, ())}
    return Facultad(aulas=("Aula 1", "Lab"), dias=("Lunes", "Martes"), turnos=2, grupos=tuple(g), anios=anios)


def test_celda_ocupacion_es_formula_de_grupos():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    # alguna celda del bloque debe referenciar una hoja de grupo y concatenar
    formulas = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    assert any("C111" in f for f in formulas)


def test_tiene_un_bloque_por_dia():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    textos = [c.value for row in ws.iter_rows() for c in row]
    assert "Lunes" in textos and "Martes" in textos


def test_formula_usa_substitute_trim_y_refs_entrecomilladas():
    """La fórmula de ocupación usa SUBSTITUTE y TRIM, y los refs de hoja llevan apóstrofos."""
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    formulas = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert any("SUBSTITUTE" in f and "TRIM" in f for f in formulas)
    # Los ids de grupo funcionan como nombres de hoja; quote_sheetname los envuelve en apóstrofos
    assert any("'C111'" in f for f in formulas)


def test_bloques_tienen_bordes():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    # Esquina superior izquierda del primer bloque: perimetro medium
    assert ws["A1"].border.left.style == "medium"
    assert ws["A1"].border.top.style == "medium"
    # Celda de ocupacion interior: fina
    assert ws["B2"].border.left.style == "thin"


def test_columnas_ajustadas_al_contenido():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    # La columna B se ajusta a "Aula 1" (las celdas de ocupacion son formulas y se ignoran).
    # El valor exacto (8) prueba que el autoajuste corrio: difiere del ancho por defecto (13).
    assert ws.column_dimensions["B"].width == len("Aula 1") + 2


def test_celdas_de_ocupacion_tienen_ajuste_de_texto():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    # B2 = primera aula, turno 1 del primer bloque: celda de ocupacion (formula)
    assert ws["B2"].alignment.wrap_text is True


def test_inmoviliza_columna_de_turnos():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    # Solo se congela la columna A (etiquetas); los bloques se apilan en vertical
    assert ws.freeze_panes == "B1"


def test_encabezados_de_bloque_con_estilo():
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    # Nombre de dia (A1) y nombre de aula (B1) del primer bloque en negrita
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    # Etiqueta de turno (A2) en negrita
    assert ws["A2"].font.bold is True


def test_hoja_aulas_tiene_leyenda_de_anios_y_conflicto():
    fac = _fac()   # facultad con año C1
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    textos = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Leyenda" in t for t in textos)
    assert any("Conflicto" in t for t in textos)
    assert any("C1" in t for t in textos)   # año presente


def test_regla_mix_emitida_para_celda():
    """El formato condicional de al menos una celda incluye una regla con 'MIX'."""
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    all_formulas = []
    for sqref, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if hasattr(rule, "formula") and rule.formula:
                all_formulas.extend(rule.formula)
    assert any("MIX" in f for f in all_formulas)


def test_reglas_color_solo_para_anios_presentes():
    """Solo se generan reglas de color para los años de la facultad, no para todos los 12."""
    fac = _fac()
    wb = Workbook()
    firmas = construir_hoja_datos(wb, fac)
    construir_hoja_aulas(wb, fac, firmas)
    ws = wb[NOMBRE_HOJA]
    all_formulas = []
    for sqref, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if hasattr(rule, "formula") and rule.formula:
                all_formulas.extend(rule.formula)
    # Solo "C1" está en fac.anios — no deben aparecer códigos de otros años
    year_codes_in_rules = [
        f for f in all_formulas
        if '"M1"' in f or '"M2"' in f or '"D1"' in f or '"D2"' in f
    ]
    assert len(year_codes_in_rules) == 0, "Se emitieron reglas para años no presentes en la facultad"
