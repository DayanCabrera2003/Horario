from openpyxl.utils import quote_sheetname, absolute_coordinate
from comun import formato
from comun import leyenda
from tribunales import layout as L
from tribunales import estilos
from tribunales.modelo import Facultad

NOMBRE_HOJA = "Localizar"
COL_ROL = "B"   # columna contigua a la de momentos: muestra el rol de la persona

# Columnas B..F de la hoja de dia y el rol que representa cada una. B=estudiante,
# C=tutor, D=oponente, E=presidente, F=secretario (ver ENCABEZADOS en layout).
_COLS_ROL = (("B", "Estudiante"), ("C", "Tutor"), ("D", "Oponente"),
             ("E", "Presidente"), ("F", "Secretario"))


def construir_hoja_localizar(wb, facultad: Facultad) -> None:
    """Crea la hoja Localizar: celda de entrada global (B1) y una tabla por cada
    (dia, local), con una fila por momento. Cada fila muestra, en la columna Rol,
    en calidad de que participa la persona buscada en ese momento."""
    ws = wb.create_sheet(NOMBRE_HOJA)
    ws["A1"] = "Localizar a:"
    ws[L.LOCALIZAR_CELDA_ENTRADA] = ""   # celda de entrada global

    # Referencia absoluta de la entrada, "$B$1" (no "$B1"): la usa el formato
    # condicional de participacion.
    entrada = absolute_coordinate(L.LOCALIZAR_CELDA_ENTRADA)
    titulos = []   # coords de titulo de cada tabla (para negrita y bordes)
    alturas = 0
    for dia in facultad.dias:
        n_mom = len(dia.momentos)
        hoja_dia = quote_sheetname(dia.fecha)
        for li, local in enumerate(facultad.locales):
            f_tit = L.localizar_fila_titulo(0, alturas)
            ws[f"{L.LOCALIZAR_COL}{f_tit}"] = f"{dia.fecha} · {local.nombre}"
            ws[f"{COL_ROL}{f_tit}"] = "Rol"
            titulos.append((f_tit, n_mom))
            for mi, momento in enumerate(dia.momentos):
                f = f_tit + 1 + mi
                ws[f"{L.LOCALIZAR_COL}{f}"] = momento.id
                r = L.fila_momento(li, mi, n_mom)
                # Rol de la persona buscada en este momento: se compara la entrada
                # contra las celdas B..F de esa fila en la hoja del dia y se devuelve
                # la etiqueta de la primera que coincide (IF anidados, no IFS, por
                # compatibilidad con Calc). Vacio si no participa.
                ws[f"{COL_ROL}{f}"] = _formula_rol(entrada, hoja_dia, r)
                # Resalta el momento (y su rol) si la persona participa en la tesis
                # asignada en (dia, local, momento).
                refs = [f"{entrada}={hoja_dia}!${col}${r}" for col, _ in _COLS_ROL]
                formula = f'AND({entrada}<>"",OR({",".join(refs)}))'
                ws.conditional_formatting.add(
                    f"{L.LOCALIZAR_COL}{f}:{COL_ROL}{f}",
                    estilos.regla_formula(formula, estilos.COLOR_LOCALIZADO))
            alturas += L.localizar_altura_tabla(n_mom)

    _aplicar_presentacion(ws, titulos, alturas)


def _formula_rol(entrada: str, hoja_dia: str, fila: int) -> str:
    """IF anidados que devuelven la etiqueta de rol de la primera columna B..F de
    `hoja_dia` cuya celda (en `fila`) coincide con la entrada; vacio si ninguna."""
    formula = '""'
    for col, etiqueta in reversed(_COLS_ROL):
        formula = f'IF({entrada}={hoja_dia}!${col}${fila},"{etiqueta}",{formula})'
    return f'=IF({entrada}="","",{formula})'


def _aplicar_presentacion(ws, titulos, alturas: int) -> None:
    """Negrita en la etiqueta de entrada y en los titulos, bordes finos por
    tabla, autoajuste y leyenda con el color de localizado."""
    col = L.LOCALIZAR_COL
    interno, externo = estilos.lado_fino(), estilos.lado_medio()
    # Bordes por tabla (titulo + filas de momento), incluyendo la columna Rol.
    for f_tit, n_mom in titulos:
        formato.aplicar_borde_tabla(
            ws, f"{col}{f_tit}:{COL_ROL}{f_tit + n_mom}", interno, externo)
    # Negrita en la etiqueta de entrada, en cada titulo de tabla y en su cabecera Rol.
    coords_negrita = ["A1"]
    coords_negrita += [f"{col}{f_tit}" for f_tit, _ in titulos]
    coords_negrita += [f"{COL_ROL}{f_tit}" for f_tit, _ in titulos]
    formato.aplicar_estilo_encabezado(
        ws, coords_negrita, estilos.fuente_encabezado(),
        estilos.fill(estilos.COLOR_ENCABEZADO))
    formato.autoajustar_columnas(ws)
    # Leyenda bajo la ultima tabla.
    fila_leyenda = L.LOCALIZAR_FILA_PRIMERA_TABLA + alturas
    ws[f"{col}{fila_leyenda}"] = "Leyenda"
    leyenda.escribir_leyenda(
        ws, f"{col}{fila_leyenda + 1}",
        [(estilos.COLOR_LOCALIZADO, "Momento donde participa la persona buscada")])
