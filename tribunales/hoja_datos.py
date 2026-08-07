from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
from tribunales.modelo import Facultad

NOMBRE_HOJA = "Datos"


def _rango_nombrado(nombre: str, celda_ini: str, celda_fin: str) -> DefinedName:
    ref = (f"{quote_sheetname(NOMBRE_HOJA)}!"
           f"{absolute_coordinate(celda_ini)}:{absolute_coordinate(celda_fin)}")
    return DefinedName(nombre, attr_text=ref)


def construir_hoja_datos(wb, facultad: Facultad) -> None:
    """Crea la hoja oculta Datos: tabla tesis->tribunal (A..E) y lista de ids de
    estudiantes (columna G), cada una con su rango nombrado."""
    ws = wb.create_sheet(NOMBRE_HOJA)
    ws.sheet_state = "hidden"

    # Tabla tesis->tribunal, sin encabezado (VLOOKUP directo). Columnas:
    # A=estudiante (clave), B=tutor(es), C=oponente, D=presidente, E=secretario,
    # F=vocal. Los co-tutores se unen con " / " en una sola casilla. Una tesis
    # conjunta genera una fila por estudiante (todas con el mismo tribunal), para
    # que al elegir cualquiera de ellos el VLOOKUP encuentre su tribunal.
    fila = 0
    for t in facultad.tesis:
        tutores = " / ".join(t.tutores)
        for est in t.estudiantes:
            fila += 1
            ws[f"A{fila}"] = est
            ws[f"B{fila}"] = tutores
            ws[f"C{fila}"] = t.oponente
            ws[f"D{fila}"] = t.presidente
            ws[f"E{fila}"] = t.secretario
            ws[f"F{fila}"] = t.vocal
    if fila:
        wb.defined_names.add(_rango_nombrado("TesisTribunal", "A1", f"F{fila}"))

    # Lista de ids de estudiantes en columna H (G en blanco de separacion, porque
    # TesisTribunal ahora llega hasta la columna F con el vocal).
    for i, e in enumerate(facultad.estudiantes, start=1):
        ws[f"H{i}"] = e.id
    n_est = len(facultad.estudiantes)
    if n_est:
        wb.defined_names.add(_rango_nombrado("EstudiantesValidos", "H1", f"H{n_est}"))
