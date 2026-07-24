from openpyxl.worksheet.datavalidation import DataValidation
from comun import formato
from comun import leyenda
from tribunales import layout as L
from tribunales import estilos
from tribunales.modelo import Dia, Facultad

# Indice de columna dentro de TesisTribunal para cada columna de profesor (C..F).
# TesisTribunal: 1=estudiante, 2=tutor, 3=oponente, 4=presidente, 5=secretario.
_IDX_VLOOKUP = {"C": 2, "D": 3, "E": 4, "F": 5}


def construir_hoja_dia(ws, dia: Dia, facultad: Facultad, asignaciones=()) -> None:
    """Rellena `ws` con un bloque por local: encabezados, filas de momento,
    dropdown de estudiante en B y lookups de tribunal en C..F. Rellena la
    columna Estudiante desde `asignaciones` (las de esta fecha)."""
    n_mom = len(dia.momentos)
    # Mapa (local_id, momento_id) -> estudiante, a partir de las asignaciones de este dia.
    puestos = {(a.local, a.momento): a.estudiante
               for a in asignaciones if a.fecha == dia.fecha}

    for li, local in enumerate(facultad.locales):
        f_tit = L.fila_titulo_local(li, n_mom)
        f_enc = L.fila_encabezado_local(li, n_mom)
        ws[f"{L.COL_MOMENTO}{f_tit}"] = local.nombre
        for col, texto in zip("ABCDEF", L.ENCABEZADOS):
            ws[f"{col}{f_enc}"] = texto
        for mi, momento in enumerate(dia.momentos):
            f = L.fila_momento(li, mi, n_mom)
            ws[f"{L.COL_MOMENTO}{f}"] = momento.id
            est = puestos.get((local.id, momento.id))
            if est is not None:
                ws[f"{L.COL_ESTUDIANTE}{f}"] = est
            for col in L.COLS_PROFESOR:
                idx = _IDX_VLOOKUP[col]
                ws[f"{col}{f}"] = (
                    f'=IF(${L.COL_ESTUDIANTE}{f}="","",'
                    f'IFERROR(VLOOKUP(${L.COL_ESTUDIANTE}{f},TesisTribunal,{idx},FALSE),""))'
                )

    _aplicar_dropdown_estudiantes(ws, dia, facultad)
    _aplicar_formato_colision(ws, dia, facultad)
    _aplicar_presentacion(ws, dia, facultad)


def _aplicar_presentacion(ws, dia: Dia, facultad: Facultad) -> None:
    """Bordes por bloque, encabezados en negrita, autoajuste, congelado de la
    columna A y leyenda con el color de colision."""
    n_mom = len(dia.momentos)
    interno, externo = estilos.lado_fino(), estilos.lado_medio()
    coords_encab = []
    for li in range(len(facultad.locales)):
        f_tit = L.fila_titulo_local(li, n_mom)
        f_enc = L.fila_encabezado_local(li, n_mom)
        f_ultima = L.fila_momento(li, n_mom - 1, n_mom)
        # Borde del bloque completo (titulo + encabezado + filas de momento).
        formato.aplicar_borde_tabla(
            ws, f"{L.COL_MOMENTO}{f_tit}:{L.COL_ULTIMA}{f_ultima}", interno, externo)
        coords_encab.append(f"{L.COL_MOMENTO}{f_tit}")                 # titulo de local
        coords_encab += [f"{col}{f_enc}" for col in "ABCDEF"]         # encabezado de columnas
    formato.aplicar_estilo_encabezado(
        ws, coords_encab, estilos.fuente_encabezado(),
        estilos.fill(estilos.COLOR_ENCABEZADO))
    formato.autoajustar_columnas(ws)
    # Inmoviliza la columna A (momentos); los bloques se apilan en vertical.
    ws.freeze_panes = "B1"
    # Leyenda tras el autoajuste, dos filas bajo el ultimo bloque.
    ultimo = len(facultad.locales) - 1
    fila_leyenda = L.fila_momento(ultimo, n_mom - 1, n_mom) + 2
    ws[f"{L.COL_MOMENTO}{fila_leyenda}"] = "Leyenda"
    leyenda.escribir_leyenda(
        ws, f"{L.COL_MOMENTO}{fila_leyenda + 1}",
        [(estilos.COLOR_COLISION, "Colision: profesor en dos locales a la vez")])


def _aplicar_formato_colision(ws, dia: Dia, facultad: Facultad) -> None:
    """Resalta una celda de profesor si ese profesor aparece en 2+ locales
    distintos en el mismo momento del dia. Se cuentan locales (no ocurrencias
    brutas): un profesor con dos roles en una misma tesis no es colision."""
    n_mom = len(dia.momentos)
    n_loc = len(facultad.locales)
    for mi in range(n_mom):
        rangos_locales = [L.rango_profesores_momento(k, mi, n_mom) for k in range(n_loc)]
        for li in range(n_loc):
            f = L.fila_momento(li, mi, n_mom)
            for col in L.COLS_PROFESOR:
                celda = f"{col}{f}"
                # Cuenta en cuantos locales aparece este profesor en este momento.
                conteo = "+".join(f"IF(COUNTIF({r},{celda})>0,1,0)" for r in rangos_locales)
                formula = f'AND({celda}<>"",({conteo})>1)'
                ws.conditional_formatting.add(
                    celda, estilos.regla_formula(formula, estilos.COLOR_COLISION))


def _aplicar_dropdown_estudiantes(ws, dia: Dia, facultad: Facultad) -> None:
    # Dropdown de estudiantes en la columna B de cada fila de momento (aviso no
    # bloqueante, como en el generador de horarios).
    n_mom = len(dia.momentos)
    dv = DataValidation(type="list", formula1="EstudiantesValidos", allow_blank=True,
                        showErrorMessage=True, errorStyle="information")
    ws.add_data_validation(dv)
    celdas = [f"{L.COL_ESTUDIANTE}{L.fila_momento(li, mi, n_mom)}"
              for li in range(len(facultad.locales))
              for mi in range(n_mom)]
    dv.sqref = " ".join(celdas)
