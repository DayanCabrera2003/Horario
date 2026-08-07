"""Importa los Excel de tribunales de Carmen y produce el `tribunal.yaml` y el
`asignaciones.yaml` que consume el generador, mas un informe de revision con el
mapa id -> nombre y las incidencias detectadas.

Las columnas se leen por posicion fija (1..10), porque los encabezados traen
erratas ('Scretario', 'Oponenete') pero el orden es estable:
  1 Dia | 2 Hora | 3 Estudiante | 4 Tutor | 5 Presidente | 6 Secretario |
  7 Vocal | 8 Oponente | 9 Local | 10 Observaciones
"""
import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

from extraccion.normalizar import (
    separar_personas, normalizar_local, parsear_hora, sumar_minutos,
    quitar_acentos, normalizar_espacios,
)
from extraccion.deduplicar import Deduplicador, _generar_id

DURACION_MOMENTO_MIN = 60   # duracion supuesta de cada defensa (no viene en el Excel)


def _fecha(valor) -> str:
    """Normaliza la fecha a 'AAAA-MM-DD', o "" si no es una fecha."""
    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat()
    if isinstance(valor, datetime.date):
        return valor.isoformat()
    return ""


def leer_filas(path):
    """Genera las filas de datos (dict con las 10 columnas por posicion) de todas
    las hojas de un Excel, saltando encabezados y filas en blanco."""
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            celdas = [ws.cell(r, c).value for c in range(1, 11)]
            if all(v in (None, "") for v in celdas):
                continue
            if str(celdas[0]).strip() == "Día":     # fila de encabezado
                continue
            yield {
                "fecha": celdas[0], "hora": celdas[1], "estudiante": celdas[2],
                "tutor": celdas[3], "presidente": celdas[4], "secretario": celdas[5],
                "vocal": celdas[6], "oponente": celdas[7], "local": celdas[8],
            }


class _Importador:
    """Acumula el estado de la importacion (dedups de personas, locales, tesis,
    asignaciones e incidencias) a medida que procesa las filas."""

    def __init__(self):
        self.prof = Deduplicador()
        self.est = Deduplicador()
        self.locales = {}          # nombre canonico -> id
        self.filas = []            # tuplas (fecha, inicio, local, idxs de roles)
        self.momentos = {}         # fecha -> set de inicios
        self.incidencias = []

    def _local_id(self, nombre: str) -> str:
        if nombre not in self.locales:
            self.locales[nombre] = _generar_id(nombre, set(self.locales.values()))
        return self.locales[nombre]

    def _persona(self, dd: Deduplicador, crudo) -> object:
        """Registra una celda de rol simple (un solo profesor). Se pasa la celda
        entera al deduplicador, que ya limpia grados repetidos y anotaciones; no se
        separa por personas para no fragmentar 'Lic. Lic. Nombre'."""
        if not crudo or not str(crudo).strip():
            return None
        return dd.agregar(str(crudo))

    def procesar(self, fila: dict) -> None:
        fecha = _fecha(fila["fecha"])
        estudiantes = [self.est.agregar(p) for p in separar_personas(fila["estudiante"])]
        estudiantes = [i for i in estudiantes if i is not None]
        if not fecha or not estudiantes:
            self.incidencias.append(f"Fila ignorada (sin fecha o sin estudiante): {fila['estudiante']!r}")
            return
        tutores = [self.prof.agregar(p) for p in separar_personas(fila["tutor"])]
        tutores = [i for i in tutores if i is not None]
        presidente = self._persona(self.prof, fila["presidente"])
        secretario = self._persona(self.prof, fila["secretario"])
        vocal = self._persona(self.prof, fila["vocal"])
        oponente = self._persona(self.prof, fila["oponente"])

        inicio = parsear_hora(fila["hora"])
        local = normalizar_local(fila["local"])
        local_id = self._local_id(local) if local else ""
        if inicio:
            self.momentos.setdefault(fecha, set()).add(inicio)
        else:
            self.incidencias.append(f"Sin hora valida ({fila['hora']!r}): {fila['estudiante']!r}")
        if not local:
            self.incidencias.append(f"Sin local reconocible ({fila['local']!r}): {fila['estudiante']!r}")

        self.filas.append({
            "fecha": fecha, "inicio": inicio, "local_id": local_id,
            "estudiantes": estudiantes, "tutores": tutores, "presidente": presidente,
            "secretario": secretario, "vocal": vocal, "oponente": oponente,
        })

    def resolver(self) -> dict:
        """Traduce los indices a ids y arma las estructuras finales del YAML."""
        prof_map = self.prof.resolver()
        est_map = self.est.resolver()

        def pid(i):
            return prof_map[i]["id"] if i is not None else ""

        profesores = [{"id": g["id"], "nombre": g["nombre"], "grado": g["grado"]}
                      for g in prof_map.values()]
        estudiantes = [{"id": g["id"], "nombre": g["nombre"]} for g in est_map.values()]
        locales = [{"id": lid, "nombre": nom} for nom, lid in self.locales.items()]

        dias = []
        for fecha in sorted(self.momentos):
            momentos = [{"inicio": h, "fin": sumar_minutos(h, DURACION_MOMENTO_MIN)}
                        for h in sorted(self.momentos[fecha])]
            dias.append({"fecha": fecha, "momentos": momentos})

        tesis, asignaciones = [], []
        for f in self.filas:
            estudiantes_ids = [est_map[i]["id"] for i in f["estudiantes"]]
            tutores_ids = [prof_map[i]["id"] for i in f["tutores"]] or [""]
            t = {"estudiantes": estudiantes_ids, "tutores": tutores_ids,
                 "oponente": pid(f["oponente"]), "presidente": pid(f["presidente"]),
                 "secretario": pid(f["secretario"])}
            if f["vocal"] is not None:
                t["vocal"] = pid(f["vocal"])
            tesis.append(t)
            if f["inicio"] and f["local_id"]:
                asignaciones.append({
                    "estudiante": estudiantes_ids[0], "local": f["local_id"],
                    "fecha": f["fecha"],
                    "momento": f"{f['inicio']}-{sumar_minutos(f['inicio'], DURACION_MOMENTO_MIN)}",
                })

        return {
            "facultad": {"profesores": profesores, "estudiantes": estudiantes,
                         "locales": locales, "dias": dias, "tesis": tesis},
            "asignaciones": asignaciones,
            "revision": {"profesores": prof_map, "estudiantes": est_map,
                         "incidencias": self.incidencias},
        }


def importar(paths) -> dict:
    """Procesa uno o varios Excel y devuelve las estructuras (facultad,
    asignaciones, revision)."""
    imp = _Importador()
    for path in paths:
        for fila in leer_filas(path):
            imp.procesar(fila)
    return imp.resolver()


class _CitadorHoras(str):
    """Marca una cadena de hora/fecha para forzar su comillado en YAML (evita que
    '09:00' se reinterprete como numero sexagesimal al recargar)."""


def _citar(dumper, data):
    return dumper.represent_scalar("tag:yaml.org,2002:str", str(data), style="'")


yaml.SafeDumper.add_representer(_CitadorHoras, _citar)


def _marcar_horas(facultad: dict) -> dict:
    """Envuelve fechas y horas en _CitadorHoras para que salgan entre comillas."""
    dias = []
    for d in facultad["dias"]:
        momentos = [{"inicio": _CitadorHoras(m["inicio"]), "fin": _CitadorHoras(m["fin"])}
                    for m in d["momentos"]]
        dias.append({"fecha": _CitadorHoras(d["fecha"]), "momentos": momentos})
    return {**facultad, "dias": dias}


def escribir_yaml(datos: dict, ruta_facultad, ruta_asignaciones) -> None:
    """Escribe tribunal.yaml (estructura) y asignaciones.yaml a partir de `datos`."""
    facultad = _marcar_horas(datos["facultad"])
    Path(ruta_facultad).write_text(
        yaml.safe_dump(facultad, sort_keys=False, allow_unicode=True,
                       default_flow_style=False),
        encoding="utf-8")
    asignaciones = [{**a, "fecha": _CitadorHoras(a["fecha"]),
                     "momento": _CitadorHoras(a["momento"])}
                    for a in datos["asignaciones"]]
    Path(ruta_asignaciones).write_text(
        yaml.safe_dump(asignaciones, sort_keys=False, allow_unicode=True,
                       default_flow_style=False),
        encoding="utf-8")


def escribir_revision(datos: dict, ruta) -> None:
    """Escribe un informe Markdown con el mapa id -> nombre (y las variantes que se
    fusionaron) y la lista de incidencias, para revisar la importacion a mano."""
    rev = datos["revision"]
    lineas = ["# Revision de la importacion de tribunales", ""]

    def tabla_personas(titulo, mapa, con_grado):
        lineas.append(f"## {titulo} ({len(mapa)})")
        lineas.append("")
        cab = "| id | nombre | grado | variantes fusionadas |" if con_grado \
            else "| id | nombre | variantes fusionadas |"
        sep = "|----|--------|-------|----------------------|" if con_grado \
            else "|----|--------|----------------------|"
        lineas.extend([cab, sep])
        for g in sorted(mapa.values(), key=lambda x: x["id"]):
            variantes = "; ".join(g["variantes"]) if g["variantes"] else ""
            if con_grado:
                lineas.append(f"| {g['id']} | {g['nombre']} | {g['grado']} | {variantes} |")
            else:
                lineas.append(f"| {g['id']} | {g['nombre']} | {variantes} |")
        lineas.append("")

    tabla_personas("Profesores", rev["profesores"], con_grado=True)
    tabla_personas("Estudiantes", rev["estudiantes"], con_grado=False)

    lineas.append(f"## Locales ({len(datos['facultad']['locales'])})")
    lineas.append("")
    lineas.extend(["| id | nombre |", "|----|--------|"])
    for l in datos["facultad"]["locales"]:
        lineas.append(f"| {l['id']} | {l['nombre']} |")
    lineas.append("")

    lineas.append(f"## Incidencias ({len(rev['incidencias'])})")
    lineas.append("")
    if rev["incidencias"]:
        lineas.extend(f"- {inc}" for inc in rev["incidencias"])
    else:
        lineas.append("Ninguna.")
    lineas.append("")
    Path(ruta).write_text("\n".join(lineas), encoding="utf-8")
