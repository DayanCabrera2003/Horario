from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
from horarios import layout as L
from horarios.modelo import Facultad

NOMBRE_HOJA = "Datos"


def _formula_firma_anio(anios_por_codigo: dict, dia_idx: int, turno: int, aula: str) -> str:
    """Fórmula que devuelve el código del año único presente en (dia,turno,aula), "MIX" si hay varios, o "" si ninguno."""
    presencias, etiquetas = [], []
    for cod, grupos in anios_por_codigo.items():
        partes = [f'COUNTIF({quote_sheetname(g.id)}!{L.celda_aula(dia_idx, turno)},"{aula}")' for g in grupos]
        presencias.append(f'({"+".join(partes)})>0')
        etiquetas.append(cod)
    suma = "+".join(f"IF({p},1,0)" for p in presencias)
    # se construye anidando: elige el año cuya presencia es verdadera cuando total=1
    elegir = '""'
    for p, cod in zip(reversed(presencias), reversed(etiquetas)):
        elegir = f'IF({p},"{cod}",{elegir})'
    # código si exactamente 1 año presente, "MIX" si >1, "" si 0
    return f'=IF(({suma})=0,"",IF(({suma})=1,{elegir},"MIX"))'


def construir_hoja_datos(wb, facultad: Facultad) -> dict[tuple[str, int, str], str]:
    """Crea la hoja oculta Datos con la lista maestra de aulas y firmas de año por celda.

    Retorna un dict que mapea (dia, turno, aula) -> "Datos!<addr>" con la dirección
    calificada de la fórmula de firma de año para esa combinación.
    """
    ws = wb.create_sheet(NOMBRE_HOJA)
    ws.sheet_state = "hidden"

    # (a) Lista maestra de aulas en columna A, con rango nombrado
    for i, aula in enumerate(facultad.aulas, start=1):
        ws[f"A{i}"] = aula
    ref = (
        f"{quote_sheetname(NOMBRE_HOJA)}!"
        f"{absolute_coordinate('A1')}:"
        f"{absolute_coordinate(f'A{len(facultad.aulas)}')}"
    )
    wb.defined_names.add(DefinedName("AulasValidas", attr_text=ref))

    # (b) Firmas de año, una por (dia, turno, aula).
    # Para cada año, presencia = suma de COUNTIF sobre las celdas-aula de los grupos de ese año.
    anios_por_codigo: dict[str, list] = {}
    for g in facultad.grupos:
        anios_por_codigo.setdefault(g.anio_codigo, []).append(g)

    celdas: dict[tuple[str, int, str], str] = {}
    fila = 1
    # La columna B se deja en blanco a propósito: separa la lista de aulas (col A)
    # de las firmas (col C) para que ambas regiones sean legibles por separado.
    col_firma = "C"   # las firmas viven en columna C hacia abajo
    for dia_idx, dia in enumerate(facultad.dias):
        for turno in range(1, facultad.turnos + 1):
            for aula in facultad.aulas:
                dir_firma = f"{col_firma}{fila}"
                ws[dir_firma] = _formula_firma_anio(anios_por_codigo, dia_idx, turno, aula)
                celdas[(dia, turno, aula)] = f"{NOMBRE_HOJA}!{dir_firma}"
                fila += 1
    return celdas
