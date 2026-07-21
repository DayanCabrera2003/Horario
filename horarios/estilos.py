from openpyxl.styles import PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule

# Colores por "año" (carrera+año). Fijos en código; el usuario no los cambia.
ANIO_COLOR = {
    "C1": "BBDEFB", "C2": "90CAF9", "C3": "64B5F6", "C4": "42A5F5",
    "M1": "C8E6C9", "M2": "A5D6A7", "M3": "81C784", "M4": "66BB6A",
    "D1": "FFE0B2", "D2": "FFCC80", "D3": "FFB74D", "D4": "FFA726",
}
COLOR_CONFLICTO = "E53935"   # rojo intenso: años distintos en la misma aula/turno
COLOR_AULA_INVALIDA = "FFF176"   # amarillo
COLOR_ASIG_DESCONOCIDA = "FFB74D"  # naranja
COLOR_SOBRE_PLANIFICADA = "EF9A9A"  # rojo
COLOR_FREC_EXACTA = "A5D6A7"       # verde


def fill(hex_rgb: str) -> PatternFill:
    return PatternFill(start_color=hex_rgb, end_color=hex_rgb, fill_type="solid")


def regla_formula(formula: str, color_hex: str) -> FormulaRule:
    return FormulaRule(formula=[formula], fill=fill(color_hex))


def lado_fino() -> Side:
    """Lado de grosor fino, para el enrejado interno de las tablas."""
    return Side(style="thin")


def lado_medio() -> Side:
    """Lado de grosor medio, para el borde exterior (perimetro) de las tablas."""
    return Side(style="medium")


def borde_fino() -> Border:
    """Borde fino en las cuatro caras, para delimitar las celdas de las tablas."""
    lado = lado_fino()
    return Border(left=lado, right=lado, top=lado, bottom=lado)
