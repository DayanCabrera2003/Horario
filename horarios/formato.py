"""Utilidades para aplicar formato de presentacion (bordes, ancho de columnas)
sobre una worksheet ya construida. Separado de `estilos`, que solo define objetos
de estilo puros, y de `layout`, que solo calcula posiciones."""
from openpyxl.styles import Border, Side


def aplicar_borde(ws, rango: str, borde: Border) -> None:
    """Aplica `borde` a cada celda del `rango` (p. ej. "C3:D15").

    openpyxl no tiene un borde "de rango": el borde es un estilo por celda, asi que
    se recorre el rango y se asigna a cada una.
    """
    for fila in ws[rango]:
        for celda in fila:
            celda.border = borde


def aplicar_borde_tabla(ws, rango: str, interno: Side, externo: Side) -> None:
    """Bordea un rango rectangular: `externo` en las caras del perimetro y
    `interno` en las caras internas.

    openpyxl no tiene borde de rango; el borde es un estilo por celda. Se
    compone el Border de cada celda segun su posicion en el rango: las celdas
    del borde reciben el lado `externo` en la cara que da al exterior.
    """
    filas = ws[rango]
    n_filas = len(filas)
    n_cols = len(filas[0]) if n_filas else 0
    for i, fila in enumerate(filas):
        for j, celda in enumerate(fila):
            celda.border = Border(
                top=externo if i == 0 else interno,
                bottom=externo if i == n_filas - 1 else interno,
                left=externo if j == 0 else interno,
                right=externo if j == n_cols - 1 else interno,
            )


def aplicar_ajuste_texto(ws, rango: str, alineacion) -> None:
    """Aplica `alineacion` (con wrap_text) a cada celda del `rango`, para que
    el contenido largo salte de linea en vez de desbordarse."""
    for fila in ws[rango]:
        for celda in fila:
            celda.alignment = alineacion


def autoajustar_columnas(ws, min_ancho: int = 8, max_ancho: int = 45,
                         extra: int = 2) -> None:
    """Fija el ancho de cada columna al del texto mas largo que contiene.

    openpyxl no tiene "autofit" real (requiere renderizar), asi que se estima por
    longitud de caracteres. Las formulas ("=...") se ignoran: su texto no coincide con
    el valor que se muestra e inflaria la columna. El ancho se limita a [min_ancho, max_ancho].
    """
    largos: dict[str, int] = {}
    for fila in ws.iter_rows():
        for celda in fila:
            valor = celda.value
            if valor is None:
                continue
            if isinstance(valor, str) and valor.startswith("="):
                continue
            largo = len(str(valor))
            if largo > largos.get(celda.column_letter, 0):
                largos[celda.column_letter] = largo
    for letra, largo in largos.items():
        ws.column_dimensions[letra].width = min(max_ancho, max(min_ancho, largo + extra))
