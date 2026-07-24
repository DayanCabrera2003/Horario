from openpyxl.worksheet.datavalidation import DataValidation
from horarios import layout as L
from horarios import estilos
from comun import formato
from horarios import leyenda
from horarios.modelo import Grupo, Facultad, Horario


def construir_hoja_grupo(ws, grupo: Grupo, facultad: Facultad,
                         horario: Horario | None = None) -> None:
    """Rellena `ws` en el sitio con la hoja de un grupo: rejilla de horario, tabla de
    asignaturas con fórmulas, dropdowns de aula/asignatura y formato condicional."""
    ws.title = grupo.id
    ws["A1"] = "Grupo"
    ws[L.CELDA_GRUPO_ID] = grupo.id

    # Encabezado de días
    for i, dia in enumerate(facultad.dias):
        ws[f"{L.col_dia(i)}{L.FILA_ENCABEZADO_DIAS}"] = dia
    # Etiquetas de turno
    for t in range(1, facultad.turnos + 1):
        ws[f"A{L.fila_asig(t)}"] = f"Turno {t}"

    # Rejilla de horario (rellena si hay horario)
    if horario:
        for (dia, turno), asg in horario.celdas.items():
            dia_idx = facultad.dias.index(dia)
            ws[L.celda_asig(dia_idx, turno)] = asg.asig
            ws[L.celda_aula(dia_idx, turno)] = asg.aula

    # Tabla de asignaturas + fórmulas
    asignaturas = facultad.asignaturas_de(grupo)
    # 'rango' cubre filas de asignatura Y de aula; COUNTIF sobre él es correcto porque
    # los nombres de aula nunca coinciden con ids de asignatura.
    rango = L.rango_horario(len(facultad.dias), facultad.turnos)
    ws["I3"] = "Asignatura"
    ws["J3"], ws["K3"], ws["L3"], ws["M3"] = "Nombre", "Frec", "Asignadas", "Faltan"
    for i, a in enumerate(asignaturas):
        id_cell = L.celda_asig_tabla_id(i)
        ws[id_cell] = a.id
        ws[L.celda_asig_tabla_nombre(i)] = a.nombre
        frec_cell = L.celda_asig_tabla_frec(i)
        ws[frec_cell] = a.frecuencia
        asignadas_cell = L.celda_asig_tabla_asignadas(i)
        ws[asignadas_cell] = f"=COUNTIF({rango},{id_cell})"
        ws[L.celda_asig_tabla_faltan(i)] = f"={frec_cell}-{asignadas_cell}"

    _aplicar_dropdown_aulas(ws, facultad)
    _aplicar_dropdown_asignaturas(ws, grupo, facultad)
    _aplicar_formato_condicional(ws, grupo, facultad)
    _aplicar_bordes(ws, grupo, facultad)
    _aplicar_estilo_encabezados(ws, grupo, facultad)
    formato.autoajustar_columnas(ws)
    # La leyenda va tras el autoajuste para que sus textos largos no ensanchen
    # la columna J (que es tambien la columna "Nombre" de la tabla).
    _aplicar_leyenda(ws, grupo, facultad)
    # Inmoviliza la fila de dias y la columna de etiquetas de turno.
    ws.freeze_panes = L.celda_asig(0, 1)


def _aplicar_bordes(ws, grupo: Grupo, facultad: Facultad) -> None:
    """Bordea la rejilla de horario, las etiquetas de turno y la tabla de
    asignaturas: enrejado fino interno y perimetro medio en cada tabla."""
    n_dias, n_turnos = len(facultad.dias), facultad.turnos
    n_asig = len(facultad.asignaturas_de(grupo))
    interno, externo = estilos.lado_fino(), estilos.lado_medio()
    formato.aplicar_borde_tabla(ws, L.rango_bloque_horario(n_dias, n_turnos), interno, externo)
    formato.aplicar_borde_tabla(ws, L.rango_etiquetas_turno(n_turnos), interno, externo)
    formato.aplicar_borde_tabla(ws, L.rango_tabla_asignaturas(n_asig), interno, externo)


def _aplicar_estilo_encabezados(ws, grupo: Grupo, facultad: Facultad) -> None:
    """Negrita + relleno neutro en dias, etiquetas de turno y cabeceras de la
    tabla de asignaturas. Solo celdas con contenido (las filas de aula en la
    columna A quedan sin tocar)."""
    coords = [f"{L.col_dia(i)}{L.FILA_ENCABEZADO_DIAS}" for i in range(len(facultad.dias))]
    coords += [f"A{L.fila_asig(t)}" for t in range(1, facultad.turnos + 1)]
    coords += ["I3", "J3", "K3", "L3", "M3"]
    formato.aplicar_estilo_encabezado(
        ws, coords, estilos.fuente_encabezado(),
        estilos.fill(estilos.COLOR_ENCABEZADO))


def _aplicar_leyenda(ws, grupo: Grupo, facultad: Facultad) -> None:
    """Mini-leyenda bajo la tabla de asignaturas (columna I), con los colores
    que aparecen en esta hoja."""
    n_asig = len(facultad.asignaturas_de(grupo))
    fila = L.FILA_PRIMERA_ASIG + n_asig - 1 + 2   # dos filas bajo la tabla
    ws[f"I{fila}"] = "Leyenda"
    items = [
        (estilos.COLOR_AULA_INVALIDA, "Aula fuera del listado de la facultad"),
        (estilos.COLOR_ASIG_DESCONOCIDA, "Asignatura fuera de la tabla del grupo"),
        (estilos.COLOR_SOBRE_PLANIFICADA, "Sobre-planificada (asignadas > frecuencia)"),
        (estilos.COLOR_FREC_EXACTA, "Frecuencia exacta cumplida"),
    ]
    leyenda.escribir_leyenda(ws, f"I{fila + 1}", items)


def _aplicar_dropdown_aulas(ws, facultad: Facultad) -> None:
    # Fuente en la hoja Datos (creada por hoja_datos). Rango nombrado 'AulasValidas'.
    # OJO: openpyxl escribe formula1 verbatim en el XML; NO lleva '=' inicial o el
    # dropdown no se puebla al abrir en Excel/LibreOffice.
    # errorStyle 'information': el aviso es no bloqueante, de modo que en Calc/Excel se
    # pueden escribir aulas fuera del listado (el dropdown queda solo como ayuda). Sin
    # una accion de error explicita, Calc asume 'Stop' y rechaza los valores nuevos.
    dv = DataValidation(type="list", formula1="AulasValidas", allow_blank=True,
                        showErrorMessage=True, errorStyle="information")
    ws.add_data_validation(dv)
    dv.sqref = L.rangos_filas_aula(len(facultad.dias), facultad.turnos)


def _aplicar_dropdown_asignaturas(ws, grupo: Grupo, facultad: Facultad) -> None:
    # Fuente: los ids de la tabla de asignaturas del propio grupo (misma hoja).
    # Sin '=' inicial y con rango absoluto para que no se desplace al insertar filas.
    n_asig = len(facultad.asignaturas_de(grupo))
    rango = L.rango_ids_asignaturas_abs(n_asig)   # $I$4:$I$5
    # Mismo criterio que las aulas: aviso 'information' no bloqueante para no rechazar
    # asignaturas escritas a mano fuera de la lista del grupo.
    dv = DataValidation(type="list", formula1=rango, allow_blank=True,
                        showErrorMessage=True, errorStyle="information")
    ws.add_data_validation(dv)
    dv.sqref = L.rangos_filas_asig(len(facultad.dias), facultad.turnos)


def _aplicar_formato_condicional(ws, grupo: Grupo, facultad: Facultad) -> None:
    n_dias, n_turnos = len(facultad.dias), facultad.turnos
    n_asig = len(facultad.asignaturas_de(grupo))
    # Rango absoluto ($I$): la regla se aplica sobre un sqref multi-rango; si fuera
    # relativo, Excel lo desplazaría por fila y comprobaría el rango equivocado.
    rango_ids = L.rango_ids_asignaturas_abs(n_asig)

    # Asignatura desconocida (naranja) en filas de asignatura
    sq_asig = L.rangos_filas_asig(n_dias, n_turnos)
    primera = L.celda_asig(0, 1)
    ws.conditional_formatting.add(
        sq_asig,
        estilos.regla_formula(
            f'AND({primera}<>"",COUNTIF({rango_ids},{primera})=0)',
            estilos.COLOR_ASIG_DESCONOCIDA),
    )
    # Aula inválida (amarillo) en filas de aula
    sq_aula = L.rangos_filas_aula(n_dias, n_turnos)
    primera_aula = L.celda_aula(0, 1)
    ws.conditional_formatting.add(
        sq_aula,
        estilos.regla_formula(
            f'AND({primera_aula}<>"",COUNTIF(AulasValidas,{primera_aula})=0)',
            estilos.COLOR_AULA_INVALIDA),
    )
    # Tabla de asignaturas: sobre-planificada (rojo) y frecuencia exacta (verde)
    l_ini = L.celda_asig_tabla_asignadas(0)
    l_fin = L.celda_asig_tabla_asignadas(n_asig - 1)
    k_ini = L.celda_asig_tabla_frec(0)
    rango_asignadas = f"{l_ini}:{l_fin}"
    ws.conditional_formatting.add(
        rango_asignadas,
        estilos.regla_formula(f"{l_ini}>{k_ini}", estilos.COLOR_SOBRE_PLANIFICADA),
    )
    ws.conditional_formatting.add(
        rango_asignadas,
        estilos.regla_formula(f"{l_ini}={k_ini}", estilos.COLOR_FREC_EXACTA),
    )
