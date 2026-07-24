from openpyxl.utils import quote_sheetname, get_column_letter
from horarios import layout as L
from horarios import estilos
from comun import formato
from comun import leyenda
from horarios.modelo import Facultad

NOMBRE_HOJA = "Aulas"


def _formula_ocupacion(facultad: Facultad, dia_idx: int, turno: int, aula: str) -> str:
    """Fórmula que concatena los ids de los grupos que ocupan (dia_idx, turno, aula)."""
    partes = [
        f'IF({quote_sheetname(g.id)}!{L.celda_aula(dia_idx, turno)}="{aula}","{g.id} ","")'
        for g in facultad.grupos
    ]
    concat = "&".join(partes)
    return f'=SUBSTITUTE(TRIM({concat})," ",",")'


def _formato_por_anio(ws, celda, dir_firma: str, facultad: Facultad) -> None:
    """Adjunta reglas de formato condicional a `celda` basadas en la firma de año."""
    coord = celda.coordinate
    # Una regla por año presente en la facultad: pinta con el color de ese año
    for cod in facultad.anios:
        color = estilos.ANIO_COLOR.get(cod)
        if color is None:
            continue
        ws.conditional_formatting.add(
            coord, estilos.regla_formula(f'{dir_firma}="{cod}"', color))
    # Conflicto: firma == "MIX"
    ws.conditional_formatting.add(
        coord, estilos.regla_formula(f'{dir_firma}="MIX"', estilos.COLOR_CONFLICTO))


def construir_hoja_aulas(wb, facultad: Facultad, firmas: dict[tuple[str, int, str], str]) -> None:
    """Crea la hoja Aulas con un bloque por día (columnas=aulas, filas=turnos).

    Cada celda de datos contiene una fórmula que concatena los grupos que ocupan
    esa aula/turno/día, y recibe formato condicional por año (según la firma de Datos)
    más una regla de conflicto para el caso MIX.
    """
    ws = wb.create_sheet(NOMBRE_HOJA, index=0)
    col_fin = get_column_letter(1 + len(facultad.aulas))
    fila = 1
    for dia_idx, dia in enumerate(facultad.dias):
        fila_ini = fila
        # Encabezado del bloque
        ws.cell(row=fila, column=1, value=dia)
        for j, aula in enumerate(facultad.aulas, start=2):
            ws.cell(row=fila, column=j, value=aula)
        fila += 1
        for turno in range(1, facultad.turnos + 1):
            ws.cell(row=fila, column=1, value=f"Turno {turno}")
            for j, aula in enumerate(facultad.aulas, start=2):
                celda = ws.cell(row=fila, column=j,
                                value=_formula_ocupacion(facultad, dia_idx, turno, aula))
                _formato_por_anio(ws, celda, firmas[(dia, turno, aula)], facultad)
            fila += 1
        # Ajuste de texto en las celdas de ocupacion: si un aula/turno junta
        # varios grupos, el resultado salta de linea y la fila crece sola.
        fila_ocup_ini = fila_ini + 1  # fila_ini es el encabezado del bloque
        formato.aplicar_ajuste_texto(ws, f"B{fila_ocup_ini}:{col_fin}{fila - 1}",
                                     estilos.alineacion_ajuste())
        # Encabezado del bloque (dia + nombres de aula) y etiquetas de turno:
        # negrita + relleno neutro.
        coords_encab = [f"{get_column_letter(c)}{fila_ini}"
                        for c in range(1, 2 + len(facultad.aulas))]
        coords_encab += [f"A{fila_ini + t}" for t in range(1, facultad.turnos + 1)]
        formato.aplicar_estilo_encabezado(
            ws, coords_encab, estilos.fuente_encabezado(),
            estilos.fill(estilos.COLOR_ENCABEZADO))
        # Bordea el bloque completo (encabezado + filas de turno): enrejado fino
        # interno y perimetro medio. La fila en blanco siguiente queda fuera para
        # separar visualmente los bloques.
        formato.aplicar_borde_tabla(ws, f"A{fila_ini}:{col_fin}{fila - 1}",
                                    estilos.lado_fino(), estilos.lado_medio())
        fila += 1  # línea en blanco entre bloques
    formato.autoajustar_columnas(ws)
    # Leyenda tras el autoajuste (para que su texto largo no ensanche la columna
    # B de aulas): colores por año presentes + conflicto, bajo el ultimo bloque.
    ws[f"A{fila}"] = "Leyenda"
    items = [(estilos.ANIO_COLOR[cod], f"Año {cod}")
             for cod in facultad.anios if cod in estilos.ANIO_COLOR]
    items.append((estilos.COLOR_CONFLICTO, "Conflicto (varios años en la misma aula)"))
    leyenda.escribir_leyenda(ws, f"A{fila + 1}", items)
    # Inmoviliza la columna A (etiquetas de turno); no hay una unica fila de
    # encabezado porque los bloques-dia se apilan en vertical.
    ws.freeze_panes = "B1"
